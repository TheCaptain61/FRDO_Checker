from __future__ import annotations

import re
import shutil
import json
import os
import subprocess
import sys
import traceback
import copy
import html
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter

import tkinter as tk
from tkinter import filedialog, messagebox, ttk


APP_TITLE = "Проверка шаблонов ФРДО СПО и ПО"
SHEET_NAME = "Шаблон"
CHECKS_SHEET_NAME = "Проверки"
CURRENT_YEAR = date.today().year
SETTINGS_FILE = "frdo_checker_settings.json"


EXPECTED_HEADERS = [
    "Наименование документа",
    "Вид документа",
    "Статус документа",
    "Подтверждение утраты",
    "Подтверждение обмена",
    "Подтверждение уничтожения",
    "Уровень образования",
    "Серия документа",
    "Номер документа",
    "Дата выдачи",
    "Регистрационный номер",
    "Код профессии, специальности",
    "Наименование профессии, специальности",
    "Наименование квалификации",
    "Наименование образовательной программы",
    "Год поступления",
    "Год окончания",
    "Срок обучения, лет",
    "Фамилия получателя",
    "Имя получателя",
    "Отчество получателя",
    "Дата рождения получателя",
    "Пол получателя",
    "СНИЛС",
    "Гражданство получателя (код страны по ОКСМ)",
    "Форма обучения",
    "Форма получения образования на момент прекращения образовательных отношений",
    "Источник финансирования обучения",
    "Наличие договора о целевом обучении",
    "Номер  договора о целевом обучении",
    "Дата заключения договора о целевом обучении",
    "Наименование организации с которой заключён договор о целевом обучении",
    "ОГРН организации с которой заключён договор о целевом обучении",
    "КПП организации с которой заключён договор о целевом обучении",
    "Наименование организации работодателя",
    "ОГРН организации работодателя",
    "КПП организации работодателя",
    "Субъект федерации в котором расположена организация работодатель",
    "Наименование документа об образовании (оригинала)",
    "Серия (оригинала)",
    "Номер (оригинала)",
    "Регистрационный N (оригинала)",
    "Дата выдачи (оригинала)",
    "Фамилия получателя (оригинала)",
    "Имя получателя (оригинала)",
    "Отчество получателя (оригинала)",
]

COL = {name: i + 1 for i, name in enumerate(EXPECTED_HEADERS)}
LETTER = {i + 1: get_column_letter(i + 1) for i in range(len(EXPECTED_HEADERS))}

ALWAYS_REQUIRED = [
    "Наименование документа",
    "Вид документа",
    "Статус документа",
    "Подтверждение утраты",
    "Подтверждение обмена",
    "Подтверждение уничтожения",
    "Уровень образования",
    "Дата выдачи",
    "Регистрационный номер",
    "Код профессии, специальности",
    "Наименование профессии, специальности",
    "Наименование квалификации",
    "Наименование образовательной программы",
    "Год поступления",
    "Год окончания",
    "Срок обучения, лет",
    "Фамилия получателя",
    "Имя получателя",
    "Отчество получателя",
    "Дата рождения получателя",
    "Пол получателя",
    "Наличие договора о целевом обучении",
]

ISSUE_ERROR = "Ошибка"
ISSUE_FIXABLE = "Можно исправить"
SEVERITY_BLOCKING = "Блокирующая"
SEVERITY_WARNING = "Предупреждение"
SEVERITY_AUTOFIX = "Автоисправление"


SPO_VALIDATION_FIELDS = {
    EXPECTED_HEADERS[index]
    for index in (1, 2, 3, 4, 5, 6, 11, 22, 24, 25, 26, 27, 28, 37)
}


@dataclass
class RuleSettings:
    check_duplicates: bool = True
    check_visible_sheet: bool = True
    check_formulas: bool = True
    check_column_structure: bool = True
    check_sheet_protection: bool = True
    parallel_scan: bool = True
    autofix_sheet_protection: bool = True
    autofix_formulas: bool = True
    autofix_column_structure: bool = True
    autofix_whitespace: bool = True
    autofix_reference_format: bool = True
    autofix_dates_years: bool = True
    autofix_snils_format: bool = True
    autofix_invalid_chars: bool = True
    autofix_conditional_cleanup: bool = True
    autofix_category_split: bool = True


def settings_path() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).with_name(SETTINGS_FILE)
    return Path(__file__).with_name(SETTINGS_FILE)


def load_settings() -> RuleSettings:
    path = settings_path()
    if not path.exists():
        return RuleSettings()
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return RuleSettings()
    defaults = RuleSettings()
    values = {}
    for field_name in defaults.__dataclass_fields__:
        values[field_name] = bool(data.get(field_name, getattr(defaults, field_name)))
    return RuleSettings(**values)


def save_settings(settings: RuleSettings) -> None:
    data = {field_name: getattr(settings, field_name) for field_name in settings.__dataclass_fields__}
    settings_path().write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


SETTINGS = load_settings()


def proposed_if_enabled(value: Any, option: str) -> Any:
    return value if getattr(SETTINGS, option) else None


@dataclass
class Issue:
    file: Path
    row: int
    col: int
    field: str
    message: str
    current: Any = None
    proposed: Any = None
    profile: str = ""
    action: str = "cell"
    extra: Any = None

    @property
    def address(self) -> str:
        if self.row and self.col:
            return f"{LETTER.get(self.col, '?')}{self.row}"
        return ""

    @property
    def status(self) -> str:
        return ISSUE_FIXABLE if self.proposed is not None else ISSUE_ERROR

    @property
    def severity(self) -> str:
        if self.proposed is not None:
            return SEVERITY_AUTOFIX
        message = self.message.casefold()
        blocking_markers = [
            "обязательное",
            "не заполнено",
            "неверный",
            "дубл",
            "отсутствует в справочнике",
            "вне допустимого",
            "меньше",
            "больше текущего",
            "раньше допустимого",
            "не менее",
            "превышена длина",
        ]
        return SEVERITY_BLOCKING if any(marker in message for marker in blocking_markers) else SEVERITY_WARNING

    @property
    def hint(self) -> str:
        if self.proposed is not None:
            return "Можно применить автоматически после подтверждения."
        message = self.message.casefold()
        if "обязательное" in message or "не заполнено" in message:
            return "Заполните поле значением из инструкции или справочника листа 'Проверки'."
        if "справочник" in message:
            return "Выберите точное значение из выпадающего списка или скрытого листа 'Проверки'."
        if "снилс" in message:
            return "Проверьте 11 цифр СНИЛС и контрольную сумму; формат: 000-000-000 00."
        if "дата" in message:
            return "Введите дату в формате дд.мм.гггг и проверьте допустимый диапазон."
        if "год" in message:
            return "Введите четыре цифры года и проверьте связь года начала/окончания."
        if "дубл" in message:
            return "Проверьте, не повторяется ли тот же документ или получатель в другом файле."
        if "символ" in message:
            return "Оставьте только символы, разрешенные для этой колонки инструкцией."
        if "длина" in message:
            return "Сократите значение до лимита для этой колонки."
        if "огрн" in message:
            return "ОГРН должен содержать 13 цифр, ОГРНИП - 15 цифр."
        if "кпп" in message:
            return "КПП должен содержать 9 цифр."
        if "номер документа для изменения" in message:
            return "Оставьте поле пустым, если инструкция не требует изменения записи."
        if "формула" in message:
            formula = self.extra.get("expected_formula") if isinstance(self.extra, dict) else None
            return f"Восстановите формулу расчетной колонки. Образец: {formula}" if formula else "Восстановите формулу расчетной колонки по соседним строкам."
        if "защита листа" in message:
            return "Проверьте структуру шаблона: лист должен оставаться защищенным, если это предусмотрено исходным шаблоном."
        return "Проверьте значение по PDF-инструкции и листу 'Проверки' в шаблоне."


@dataclass(frozen=True)
class TemplateProfile:
    code: str
    title: str
    headers: list[str]
    scanner: Callable[[Path, Any], list[Issue]]


@dataclass(frozen=True)
class CategorySplit:
    parts: tuple[str, ...]

    def __str__(self) -> str:
        return "разделить на строки: " + ", ".join(self.parts)


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", text(value).replace("\u00a0", " ")).strip()


def normalize_dash(value: str) -> str:
    return value.replace("–", "-").replace("—", "-")


def parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = clean_text(value)
    if not s:
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def parse_year(value: Any) -> int | None:
    if isinstance(value, datetime):
        return value.year
    if isinstance(value, date):
        return value.year
    if isinstance(value, (int, float)) and int(value) == value:
        return int(value)
    s = clean_text(value)
    if re.fullmatch(r"\d{4}", s):
        return int(s)
    return None


def value_for_display(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    return text(value)


def extract_defined_list(workbook, name: str) -> list[Any]:
    defined = workbook.defined_names.get(name)
    if not defined:
        return []
    values: list[Any] = []
    for sheet_name, coord in defined.destinations:
        ws = workbook[sheet_name]
        for row in ws[coord]:
            cells = row if isinstance(row, tuple) else (row,)
            for cell in cells:
                if not is_blank(cell.value):
                    values.append(cell.value)
    return values


def load_reference_lists(workbook) -> dict[str, list[Any]]:
    return {
        "Вид документа": extract_defined_list(workbook, "спВиды"),
        "Статус документа": extract_defined_list(workbook, "спСтатусыДокумента"),
        "Подтверждение": extract_defined_list(workbook, "спПодтверждение"),
        "Уровень образования": extract_defined_list(workbook, "СпУровниОбразования"),
        "Код профессии, специальности": extract_defined_list(workbook, "списокКодов"),
        "Пол получателя": extract_defined_list(workbook, "спПол"),
        "Гражданство получателя (код страны по ОКСМ)": extract_defined_list(workbook, "ИН"),
        "Форма обучения": extract_defined_list(workbook, "ФОРМА"),
        "Форма получения образования на момент прекращения образовательных отношений": extract_defined_list(workbook, "ОТ"),
        "Источник финансирования обучения": extract_defined_list(workbook, "ФИН"),
        "Наличие договора о целевом обучении": extract_defined_list(workbook, "дог"),
        "Субъект федерации в котором расположена организация работодатель": extract_defined_list(workbook, "Рег"),
    }


def normalized_key(value: Any) -> str:
    s = normalize_dash(clean_text(value)).casefold()
    if re.fullmatch(r"\d+(\.0)?", s):
        return str(int(float(s)))
    return s


def list_match(value: Any, allowed: list[Any]) -> tuple[bool, Any | None]:
    if not allowed or is_blank(value):
        return True, None
    current_key = normalized_key(value)
    allowed_keys = {normalized_key(v): v for v in allowed}
    if current_key in allowed_keys:
        exact = allowed_keys[current_key]
        if text(value) != text(exact):
            return True, exact
        return True, None
    return False, None


def parse_category_list(value: Any) -> list[str]:
    raw = clean_text(value).upper()
    if not raw:
        return []
    return [part.strip() for part in re.split(r"[,;]+", raw) if part.strip()]


def valid_category_list(value: Any, allowed: list[Any]) -> tuple[bool, CategorySplit | None]:
    parts = parse_category_list(value)
    if len(parts) < 2:
        return False, None
    allowed_keys = {normalized_key(v).upper() for v in allowed}
    if all(part in allowed_keys for part in parts):
        return True, CategorySplit(tuple(parts))
    return False, None


def split_category_list(value: Any, allowed: list[Any]) -> tuple[bool, str | None, str | None]:
    parts = parse_category_list(value)
    if len(parts) < 2 or "D" not in parts:
        return False, None, None
    allowed_keys = {normalized_key(v).upper() for v in allowed}
    unique_parts = list(dict.fromkeys(parts))
    if not all(part in allowed_keys for part in unique_parts):
        return False, None, None
    keep_parts = [part for part in unique_parts if part != "D"]
    if not keep_parts:
        return False, None, None
    return True, ", ".join(keep_parts), "D"


def valid_chars(pattern: str, value: Any) -> bool:
    s = clean_text(value)
    return not s or re.fullmatch(pattern, s, flags=re.IGNORECASE) is not None


LATIN_TO_CYRILLIC_LOOKALIKE = str.maketrans(
    {
        "A": "А",
        "B": "В",
        "C": "С",
        "E": "Е",
        "H": "Н",
        "K": "К",
        "M": "М",
        "O": "О",
        "P": "Р",
        "T": "Т",
        "X": "Х",
        "Y": "У",
        "a": "а",
        "c": "с",
        "e": "е",
        "o": "о",
        "p": "р",
        "x": "х",
        "y": "у",
    }
)

CYRILLIC = "АБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯабвгдеёжзийклмнопрстуфхцчшщъыьэюя"
ASCII_LETTERS = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz"
DIGITS = "0123456789"


def normalize_common_symbols(value: Any) -> str:
    return (
        normalize_dash(clean_text(value))
        .replace("`", "'")
        .replace("´", "'")
        .replace("’", "'")
        .replace("‘", "'")
        .replace("ʼ", "'")
        .replace("“", '"')
        .replace("”", '"')
        .replace("„", '"')
        .replace("«", '"')
        .replace("»", '"')
        .replace("№", "N")
    )


def sanitize_allowed(value: Any, allowed: str, *, transliterate_lookalikes: bool = False, strip_edges: str = " ") -> str | None:
    s = normalize_common_symbols(value)
    if transliterate_lookalikes:
        s = s.translate(LATIN_TO_CYRILLIC_LOOKALIKE)
    cleaned = "".join(ch for ch in s if ch in allowed)
    cleaned = re.sub(r"\s+", " ", cleaned)
    cleaned = re.sub(r"-{2,}", "-", cleaned)
    cleaned = cleaned.strip(strip_edges)
    return cleaned if cleaned and cleaned != text(value) else None


def sanitize_name(value: Any) -> str | None:
    return sanitize_allowed(value, CYRILLIC + " .'-()", transliterate_lookalikes=True, strip_edges=" -")


def sanitize_doc_name(value: Any) -> str | None:
    return sanitize_allowed(value, CYRILLIC + " -", transliterate_lookalikes=True, strip_edges=" -")


def sanitize_series(value: Any) -> str | None:
    return sanitize_allowed(value, CYRILLIC + ASCII_LETTERS + DIGITS + " .-/", strip_edges=" ")


def sanitize_digits(value: Any) -> str | None:
    cleaned = re.sub(r"\D", "", text(value))
    return cleaned if cleaned and cleaned != text(value) else None


def sanitize_spo_registration(value: Any) -> str | None:
    return sanitize_allowed(value, CYRILLIC + ASCII_LETTERS + DIGITS + "()./- ", strip_edges=" ")


def sanitize_po_registration(value: Any) -> str | None:
    return sanitize_allowed(value, CYRILLIC + ASCII_LETTERS + DIGITS + "N()./-_ ", strip_edges=" ")


def sanitize_common_name(value: Any) -> str | None:
    return sanitize_allowed(value, CYRILLIC + ASCII_LETTERS + DIGITS + "()./-:;, ", strip_edges=" ")


def sanitize_po_program(value: Any) -> str | None:
    return sanitize_allowed(value, CYRILLIC + ASCII_LETTERS + DIGITS + 'N"().,?:/- &_#+;', strip_edges=" ")


def has_bad_edges_or_repeats(value: Any) -> bool:
    s = text(value)
    return (
        s.startswith((" ", "-"))
        or s.endswith((" ", "-"))
        or "  " in s
        or "--" in s
    )


def snils_checksum_ok(snils_digits: str) -> bool:
    if not re.fullmatch(r"\d{11}", snils_digits):
        return False
    number = snils_digits[:9]
    checksum = int(snils_digits[9:])
    total = sum(int(number[i]) * (9 - i) for i in range(9))
    if total < 100:
        expected = total
    elif total in (100, 101):
        expected = 0
    else:
        expected = total % 101
        if expected == 100:
            expected = 0
    return checksum == expected


def format_snils(snils_digits: str) -> str:
    return f"{snils_digits[:3]}-{snils_digits[3:6]}-{snils_digits[6:9]} {snils_digits[9:]}"


def row_has_data(ws, row: int) -> bool:
    # Templates usually contain formulas in R down to the last prepared row.
    # Conditional target-contract fields can also be prefilled with placeholders,
    # so a real record is detected by core document/person/original fields.
    trigger_cols = list(range(COL["Серия документа"], COL["Источник финансирования обучения"] + 1))
    trigger_cols.remove(COL["Срок обучения, лет"])
    trigger_cols.extend(range(COL["Наименование документа об образовании (оригинала)"], COL["Отчество получателя (оригинала)"] + 1))
    return any(not is_blank(ws.cell(row, col).value) for col in trigger_cols)


def add_issue(issues: list[Issue], file: Path, row: int, col: int, message: str, current: Any = None, proposed: Any = None) -> None:
    issues.append(Issue(file, row, col, EXPECTED_HEADERS[col - 1] if col else "", message, current, proposed, "СПО"))


def check_required(issues: list[Issue], file: Path, ws, row: int, field: str) -> None:
    col = COL[field]
    if is_blank(ws.cell(row, col).value):
        add_issue(issues, file, row, col, "Обязательное поле не заполнено")


def maybe_trim(issues: list[Issue], file: Path, ws, row: int, col: int) -> str:
    value = ws.cell(row, col).value
    cleaned = normalize_dash(clean_text(value))
    if isinstance(value, str) and value != cleaned:
        add_issue(issues, file, row, col, "Лишние пробелы или длинное тире", value, proposed_if_enabled(cleaned, "autofix_whitespace"))
    return cleaned


def check_length(issues: list[Issue], file: Path, ws, row: int, field: str, max_len: int) -> None:
    col = COL[field]
    if len(clean_text(ws.cell(row, col).value)) > max_len:
        add_issue(issues, file, row, col, f"Превышена длина: максимум {max_len} символов", ws.cell(row, col).value)


def check_regex(
    issues: list[Issue],
    file: Path,
    ws,
    row: int,
    field: str,
    pattern: str,
    description: str,
    sanitizer: Callable[[Any], str | None] | None = None,
) -> None:
    col = COL[field]
    value = ws.cell(row, col).value
    if not is_blank(value) and not valid_chars(pattern, value):
        proposed = sanitizer(value) if sanitizer and SETTINGS.autofix_invalid_chars else None
        if proposed and valid_chars(pattern, proposed):
            add_issue(issues, file, row, col, description, value, proposed)
        else:
            add_issue(issues, file, row, col, description, value)


def is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.strip().startswith("=")


def translate_formula(formula: str, origin_row: int, origin_col: int, target_row: int, target_col: int) -> str | None:
    try:
        origin = f"{get_column_letter(origin_col)}{origin_row}"
        target = f"{get_column_letter(target_col)}{target_row}"
        return Translator(formula, origin=origin).translate_formula(target)
    except Exception:
        return None


def nearest_formula(ws, row: int, col: int, data_rows: list[int]) -> tuple[int, str] | None:
    formula_rows = [
        candidate_row
        for candidate_row in data_rows
        if candidate_row != row and is_formula(ws.cell(candidate_row, col).value)
    ]
    if not formula_rows:
        return None
    source_row = min(formula_rows, key=lambda candidate_row: abs(candidate_row - row))
    return source_row, text(ws.cell(source_row, col).value)


def check_formula_columns(
    issues: list[Issue],
    path: Path,
    ws,
    headers: list[str],
    has_data: Callable[[Any, int], bool],
    profile: str,
) -> None:
    if not SETTINGS.check_formulas:
        return
    max_col = len(headers)
    data_rows = [row for row in range(2, ws.max_row + 1) if has_data(ws, row)]
    if not data_rows:
        return
    sample_rows = data_rows[: min(25, len(data_rows))]
    formula_cols: dict[int, str] = {}
    for col in range(1, max_col + 1):
        samples = [ws.cell(row, col).value for row in sample_rows]
        formula_count = sum(1 for value in samples if is_formula(value))
        if formula_count:
            first_formula = next(text(value) for value in samples if is_formula(value))
            formula_cols[col] = first_formula

    for col, expected_formula in formula_cols.items():
        field = headers[col - 1]
        for row in data_rows:
            value = ws.cell(row, col).value
            if not is_formula(value):
                source = nearest_formula(ws, row, col, data_rows)
                proposed = None
                if source and SETTINGS.autofix_formulas:
                    source_row, source_formula = source
                    proposed = translate_formula(source_formula, source_row, col, row, col)
                issues.append(
                    Issue(
                        path,
                        row,
                        col,
                        field,
                        "Формула в расчетной колонке отсутствует или заменена значением",
                        value,
                        proposed,
                        profile,
                        extra={"expected_formula": expected_formula},
                    )
                )


def validation_columns(ws, max_col: int) -> set[int]:
    columns: set[int] = set()
    for validation in ws.data_validations.dataValidation:
        for cell_range in validation.cells.ranges:
            min_col, min_row, max_range_col, max_row = cell_range.bounds
            if max_row < 2:
                continue
            for col in range(min_col, min(max_range_col, max_col) + 1):
                columns.add(col)
    return columns


def cell_has_validation(ws, row: int, col: int) -> bool:
    for validation in ws.data_validations.dataValidation:
        for cell_range in validation.cells.ranges:
            min_col, min_row, max_col, max_row = cell_range.bounds
            if min_row <= row <= max_row and min_col <= col <= max_col:
                return True
    return False


def check_column_structure(
    issues: list[Issue],
    path: Path,
    ws,
    headers: list[str],
    has_data: Callable[[Any, int], bool],
    profile: str,
    expected_validation_fields: set[str] | None = None,
) -> None:
    if not SETTINGS.check_column_structure:
        return
    expected_max_col = len(headers)
    for col in range(expected_max_col + 1, ws.max_column + 1):
        values = [ws.cell(row, col).value for row in range(1, min(ws.max_row, 50) + 1)]
        if any(not is_blank(value) for value in values):
            issues.append(
                Issue(
                    path,
                    1,
                    col,
                    f"Лишняя колонка {get_column_letter(col)}",
                    "В файле есть данные за пределами структуры шаблона",
                    value_for_display(next((value for value in values if not is_blank(value)), "")),
                    profile=profile,
                )
            )

    data_rows = [row for row in range(2, ws.max_row + 1) if has_data(ws, row)]
    checked_rows = data_rows[:50]
    validation_cols = validation_columns(ws, expected_max_col)
    validation_field_keys = {
        normalized_key(field)
        for field in (expected_validation_fields or set())
    }
    for col, field in enumerate(headers, start=1):
        dimension = ws.column_dimensions[get_column_letter(col)]
        if dimension.hidden:
            issues.append(
                Issue(
                    path,
                    1,
                    col,
                    field,
                    "Колонка шаблона скрыта",
                    "скрыта",
                    proposed_if_enabled("показать колонку", "autofix_column_structure"),
                    profile,
                    "unhide_column",
                )
            )
        width = dimension.width
        if width is not None and (width < 3 or width > 80):
            issues.append(Issue(path, 1, col, field, f"Подозрительная ширина колонки: {width}", width, profile=profile))
        if normalized_key(field) in validation_field_keys and col in validation_cols and checked_rows:
            missing_validation_rows = [
                row
                for row in checked_rows
                if not cell_has_validation(ws, row, col)
            ]
            if missing_validation_rows:
                issues.append(
                    Issue(
                        path,
                        missing_validation_rows[0],
                        col,
                        field,
                        "Проверка данных/выпадающий список отличается между строками",
                        f"строк без проверки: {len(missing_validation_rows)}",
                        profile=profile,
                    )
                )


def check_sheet_protection(issues: list[Issue], path: Path, ws, profile: str) -> None:
    if SETTINGS.check_sheet_protection and not ws.protection.sheet:
        issues.append(
            Issue(
                path,
                0,
                0,
                "",
                "Защита листа 'Шаблон' снята",
                current="защита отключена",
                proposed=proposed_if_enabled("включить защиту листа", "autofix_sheet_protection"),
                profile=profile,
                action="protect_sheet",
            )
        )


def scan_spo_workbook(path: Path, workbook=None) -> list[Issue]:
    issues: list[Issue] = []
    own_workbook = workbook is None
    if own_workbook:
        workbook = load_workbook(path, read_only=False, data_only=False)

    if SHEET_NAME not in workbook.sheetnames:
        issues.append(Issue(path, 0, 0, "", f"Нет листа '{SHEET_NAME}'", profile="СПО"))
        return issues

    visible_sheets = [ws.title for ws in workbook.worksheets if ws.sheet_state == "visible"]
    if SETTINGS.check_visible_sheet and visible_sheets != [SHEET_NAME]:
        issues.append(Issue(path, 0, 0, "", "В файле должен быть один видимый лист с данными: 'Шаблон'", profile="СПО"))

    ws = workbook[SHEET_NAME]
    refs = load_reference_lists(workbook)
    check_sheet_protection(issues, path, ws, "СПО")
    check_column_structure(issues, path, ws, EXPECTED_HEADERS, row_has_data, "СПО", SPO_VALIDATION_FIELDS)

    headers = [clean_text(ws.cell(1, col).value) for col in range(1, len(EXPECTED_HEADERS) + 1)]
    for col, (actual, expected) in enumerate(zip(headers, EXPECTED_HEADERS), start=1):
        if normalized_key(actual) != normalized_key(expected):
            add_issue(issues, path, 1, col, f"Неверный заголовок, ожидается: {expected}", actual, expected)

    for row in range(2, ws.max_row + 1):
        if not row_has_data(ws, row):
            continue

        for col in range(1, len(EXPECTED_HEADERS) + 1):
            maybe_trim(issues, path, ws, row, col)

        for field in ALWAYS_REQUIRED:
            check_required(issues, path, ws, row, field)

        doc_name = clean_text(ws.cell(row, COL["Наименование документа"]).value)
        status = clean_text(ws.cell(row, COL["Статус документа"]).value)
        issue_date = parse_date(ws.cell(row, COL["Дата выдачи"]).value)
        targeted = clean_text(ws.cell(row, COL["Наличие договора о целевом обучении"]).value)

        if normalized_key(doc_name) != normalized_key("Справка об обучении"):
            series_col = COL["Серия документа"]
            if is_blank(ws.cell(row, series_col).value):
                add_issue(issues, path, row, series_col, "Серия документа не заполнена; при отсутствии серии по инструкции указывается 'Нет'", ws.cell(row, series_col).value, proposed_if_enabled("Нет", "autofix_reference_format"))
            check_required(issues, path, ws, row, "Номер документа")

        if issue_date and issue_date.year >= 2021:
            for field in [
                "СНИЛС",
                "Гражданство получателя (код страны по ОКСМ)",
                "Форма обучения",
                "Форма получения образования на момент прекращения образовательных отношений",
                "Источник финансирования обучения",
            ]:
                check_required(issues, path, ws, row, field)

        suppressed_cols: set[int] = set()

        if normalized_key(targeted) == normalized_key("Да"):
            for field in EXPECTED_HEADERS[29:38]:
                check_required(issues, path, ws, row, field)
        elif normalized_key(targeted) == normalized_key("Нет"):
            for field in EXPECTED_HEADERS[29:38]:
                col = COL[field]
                if not is_blank(ws.cell(row, col).value):
                    add_issue(issues, path, row, col, "Поле заполняется только при договоре о целевом обучении = Да", ws.cell(row, col).value, proposed_if_enabled("", "autofix_conditional_cleanup"))
                    suppressed_cols.add(col)

        if normalized_key(status) == normalized_key("Дубликат"):
            for field in EXPECTED_HEADERS[38:46]:
                check_required(issues, path, ws, row, field)
        elif normalized_key(status) == normalized_key("Оригинал"):
            for field in EXPECTED_HEADERS[38:46]:
                col = COL[field]
                if not is_blank(ws.cell(row, col).value):
                    add_issue(issues, path, row, col, "Сведения об оригинале заполняются только для статуса 'Дубликат'", ws.cell(row, col).value, proposed_if_enabled("", "autofix_conditional_cleanup"))
                    suppressed_cols.add(col)

        list_fields = {
            "Вид документа": "Вид документа",
            "Статус документа": "Статус документа",
            "Подтверждение утраты": "Подтверждение",
            "Подтверждение обмена": "Подтверждение",
            "Подтверждение уничтожения": "Подтверждение",
            "Уровень образования": "Уровень образования",
            "Код профессии, специальности": "Код профессии, специальности",
            "Пол получателя": "Пол получателя",
            "Гражданство получателя (код страны по ОКСМ)": "Гражданство получателя (код страны по ОКСМ)",
            "Форма обучения": "Форма обучения",
            "Форма получения образования на момент прекращения образовательных отношений": "Форма получения образования на момент прекращения образовательных отношений",
            "Источник финансирования обучения": "Источник финансирования обучения",
            "Наличие договора о целевом обучении": "Наличие договора о целевом обучении",
            "Субъект федерации в котором расположена организация работодатель": "Субъект федерации в котором расположена организация работодатель",
        }
        for field, ref_name in list_fields.items():
            col = COL[field]
            if col in suppressed_cols:
                continue
            ok, proposed = list_match(ws.cell(row, col).value, refs.get(ref_name, []))
            if proposed is not None:
                add_issue(issues, path, row, col, "Значение отличается от справочника только форматом", ws.cell(row, col).value, proposed_if_enabled(proposed, "autofix_reference_format"))
            elif not ok:
                add_issue(issues, path, row, col, "Значение отсутствует в справочнике", ws.cell(row, col).value)

        date_fields = ["Дата выдачи", "Дата рождения получателя", "Дата заключения договора о целевом обучении", "Дата выдачи (оригинала)"]
        for field in date_fields:
            col = COL[field]
            if col in suppressed_cols:
                continue
            value = ws.cell(row, col).value
            if is_blank(value):
                continue
            parsed = parse_date(value)
            if not parsed:
                add_issue(issues, path, row, col, "Неверный формат даты, нужен дд.мм.гггг", value)
            elif not isinstance(value, (datetime, date)) and clean_text(value) != parsed.strftime("%d.%m.%Y"):
                add_issue(issues, path, row, col, "Дата будет приведена к формату дд.мм.гггг", value, proposed_if_enabled(parsed, "autofix_dates_years"))

        if issue_date and not (date(1950, 1, 1) <= issue_date <= date(2100, 12, 31)):
            add_issue(issues, path, row, COL["Дата выдачи"], "Дата выдачи вне допустимого диапазона 1950-2100", ws.cell(row, COL["Дата выдачи"]).value)

        birth_date = parse_date(ws.cell(row, COL["Дата рождения получателя"]).value)
        if birth_date and not (date(1900, 1, 1) <= birth_date <= date(2100, 12, 31)):
            add_issue(issues, path, row, COL["Дата рождения получателя"], "Дата рождения вне допустимого диапазона 1900-2100", ws.cell(row, COL["Дата рождения получателя"]).value)

        start_year = parse_year(ws.cell(row, COL["Год поступления"]).value)
        end_year = parse_year(ws.cell(row, COL["Год окончания"]).value)
        min_start = 1955 if normalized_key(status) == normalized_key("Дубликат") else 1978
        for field, year in [("Год поступления", start_year), ("Год окончания", end_year)]:
            col = COL[field]
            raw = ws.cell(row, col).value
            if is_blank(raw):
                continue
            if year is None:
                add_issue(issues, path, row, col, "Нужен год в формате гггг", raw)
            elif text(raw) != str(year):
                add_issue(issues, path, row, col, "Год будет приведён к формату гггг", raw, proposed_if_enabled(year, "autofix_dates_years"))
        if start_year:
            if start_year < min_start:
                add_issue(issues, path, row, COL["Год поступления"], f"Год поступления раньше допустимого минимума {min_start}", start_year)
            if start_year > CURRENT_YEAR:
                add_issue(issues, path, row, COL["Год поступления"], f"Год поступления больше текущего года {CURRENT_YEAR}", start_year)
        if end_year:
            if end_year > CURRENT_YEAR:
                add_issue(issues, path, row, COL["Год окончания"], f"Год окончания больше текущего года {CURRENT_YEAR}", end_year)
            if start_year and end_year < start_year:
                add_issue(issues, path, row, COL["Год окончания"], "Год окончания меньше года поступления", end_year)
        if start_year is not None and end_year is not None and end_year >= start_year:
            expected_term = end_year - start_year
            term_col = COL["Срок обучения, лет"]
            raw_term = ws.cell(row, term_col).value
            actual_term = parse_year(raw_term)
            if isinstance(raw_term, str) and raw_term.strip().startswith("="):
                pass
            elif actual_term != expected_term:
                add_issue(issues, path, row, term_col, "Срок обучения должен равняться: год окончания минус год поступления", ws.cell(row, term_col).value, proposed_if_enabled(expected_term, "autofix_dates_years"))

        length_limits = {
            "Наименование документа": 20,
            "Серия документа": 20,
            "Номер документа": 30,
            "Регистрационный номер": 30,
            "Наименование профессии, специальности": 255,
            "Наименование квалификации": 255,
            "Наименование образовательной программы": 255,
            "Фамилия получателя": 50,
            "Имя получателя": 50,
            "Отчество получателя": 50,
            "Номер  договора о целевом обучении": 30,
            "Наименование организации с которой заключён договор о целевом обучении": 200,
            "Наименование организации работодателя": 200,
            "Наименование документа об образовании (оригинала)": 100,
            "Серия (оригинала)": 20,
            "Номер (оригинала)": 20,
            "Регистрационный N (оригинала)": 50,
            "Фамилия получателя (оригинала)": 50,
            "Имя получателя (оригинала)": 50,
            "Отчество получателя (оригинала)": 50,
        }
        for field, limit in length_limits.items():
            check_length(issues, path, ws, row, field, limit)

        name_pattern = r"[А-ЯЁа-яё .'\-()]+"
        for field in ["Фамилия получателя", "Имя получателя", "Отчество получателя", "Фамилия получателя (оригинала)", "Имя получателя (оригинала)", "Отчество получателя (оригинала)"]:
            col = COL[field]
            if col in suppressed_cols:
                continue
            value = ws.cell(row, col).value
            if not is_blank(value):
                if not valid_chars(name_pattern, value):
                    add_issue(issues, path, row, col, "Недопустимые символы: разрешены кириллица, пробел, дефис, точка, апостроф, скобки", value, sanitize_name(value))
                if has_bad_edges_or_repeats(value):
                    add_issue(issues, path, row, col, "Пробел/дефис в начале или конце либо повтор пробелов/дефисов", value, clean_text(value).strip("- "))

        check_regex(issues, path, ws, row, "Наименование документа", r"[А-ЯЁа-яё \-]+", "Недопустимые символы: разрешены кириллица, пробел, дефис", sanitize_doc_name)
        for field in ["Серия документа", "Серия (оригинала)"]:
            check_regex(issues, path, ws, row, field, r"[А-ЯЁа-яёA-Z0-9 .\-/]+", "Недопустимые символы в серии", sanitize_series)
        for field in ["Номер документа", "Номер (оригинала)"]:
            check_regex(issues, path, ws, row, field, r"\d+", "Номер должен содержать только цифры", sanitize_digits)
        for field in ["Регистрационный номер", "Регистрационный N (оригинала)"]:
            check_regex(issues, path, ws, row, field, r"[А-ЯЁа-яёA-Z0-9()./\- ]+", "Недопустимые символы в регистрационном номере", sanitize_spo_registration)
        for field in ["Наименование профессии, специальности", "Наименование квалификации", "Наименование образовательной программы"]:
            check_regex(issues, path, ws, row, field, r"[А-ЯЁа-яёA-Z0-9()./\-:;, ]+", "Недопустимые символы в наименовании", sanitize_common_name)

        snils_col = COL["СНИЛС"]
        snils = clean_text(ws.cell(row, snils_col).value)
        if snils:
            digits = re.sub(r"\D", "", snils)
            if len(digits) == 11 and snils_checksum_ok(digits):
                formatted = format_snils(digits)
                if snils != formatted:
                    add_issue(issues, path, row, snils_col, "СНИЛС будет приведён к формату 000-000-000 00", ws.cell(row, snils_col).value, proposed_if_enabled(formatted, "autofix_snils_format"))
            else:
                add_issue(issues, path, row, snils_col, "Неверный СНИЛС: нужен формат 000-000-000 00 и корректная контрольная сумма", ws.cell(row, snils_col).value)

        for field in ["ОГРН организации с которой заключён договор о целевом обучении", "ОГРН организации работодателя"]:
            col = COL[field]
            if col in suppressed_cols:
                continue
            value = clean_text(ws.cell(row, col).value)
            if value and not re.fullmatch(r"\d{13}|\d{15}", value):
                add_issue(issues, path, row, col, "ОГРН должен содержать 13 или 15 цифр", ws.cell(row, col).value)
        for field in ["КПП организации с которой заключён договор о целевом обучении", "КПП организации работодателя"]:
            col = COL[field]
            if col in suppressed_cols:
                continue
            value = clean_text(ws.cell(row, col).value)
            if value and not re.fullmatch(r"\d{9}", value):
                add_issue(issues, path, row, col, "КПП должен содержать 9 цифр", ws.cell(row, col).value)

    check_formula_columns(issues, path, ws, EXPECTED_HEADERS, row_has_data, "СПО")

    if own_workbook:
        workbook.close()
    return issues


PO_HEADERS = [
    "Вид документа",
    "Статус документа",
    "Подтверждение утраты",
    "Подтверждение обмена",
    "Подтверждение уничтожения",
    "Серия документа",
    "Номер документа",
    "Дата выдачи документа",
    "Регистрационный номер",
    "Программа профессионального обучения, направление подготовки",
    "Наименование программы профессионального обучения",
    "Наименование профессий рабочих, должностей служащих",
    "Присвоенный квалификационный разряд, класс, категория (при наличии)",
    "Год начала обучения",
    "Год окончания обучения",
    "Срок обучения, часов",
    "Фамилия получателя",
    "Имя получателя",
    "Отчество получателя",
    "Дата рождения получателя",
    "Пол получателя",
    "СНИЛС",
    "Гражданство получателя (код страны по ОКСМ)",
    "Форма обучения",
    "Источник финансирования обучения",
    "Форма получения образования на момент прекращения образовательных отношений",
    "Наименование документа об образовании (оригинала)",
    "Серия (оригинала)",
    "Номер (оригинала)",
    "Регистрационный N (оригинала)",
    "Дата выдачи (оригинала)",
    "Фамилия получателя (оригинала)",
    "Имя получателя (оригинала)",
    "Отчество получателя (оригинала)",
    "Номер документа для изменения",
]


def po_col(field: str) -> int:
    return PO_HEADERS.index(field) + 1


def po_row_has_data(ws, row: int) -> bool:
    trigger_cols = list(range(po_col("Серия документа"), po_col("Форма получения образования на момент прекращения образовательных отношений") + 1))
    trigger_cols.extend(range(po_col("Наименование документа об образовании (оригинала)"), po_col("Отчество получателя (оригинала)") + 1))
    return any(not is_blank(ws.cell(row, col).value) for col in trigger_cols)


def po_add_issue(issues: list[Issue], file: Path, row: int, col: int, message: str, current: Any = None, proposed: Any = None) -> None:
    field = PO_HEADERS[col - 1] if col else ""
    issues.append(Issue(file, row, col, field, message, current, proposed, "ПО"))


def po_add_split_issue(issues: list[Issue], file: Path, row: int, col: int, current: Any, keep_value: str, duplicate_value: str) -> None:
    field = PO_HEADERS[col - 1] if col else ""
    issues.append(
        Issue(
            file,
            row,
            col,
            field,
            "Строка будет разделена по категориям: исходная строка + строка-дубликат",
            current,
            f"{keep_value}; новая строка: {duplicate_value}",
            "ПО",
            "split_row",
            {"keep_value": keep_value, "duplicate_value": duplicate_value},
        )
    )


def po_check_required(issues: list[Issue], file: Path, ws, row: int, field: str) -> None:
    col = po_col(field)
    if is_blank(ws.cell(row, col).value):
        po_add_issue(issues, file, row, col, "Обязательное поле не заполнено")


def po_check_length(issues: list[Issue], file: Path, ws, row: int, field: str, max_len: int) -> None:
    col = po_col(field)
    if len(clean_text(ws.cell(row, col).value)) > max_len:
        po_add_issue(issues, file, row, col, f"Превышена длина: максимум {max_len} символов", ws.cell(row, col).value)


def po_check_regex(
    issues: list[Issue],
    file: Path,
    ws,
    row: int,
    field: str,
    pattern: str,
    description: str,
    sanitizer: Callable[[Any], str | None] | None = None,
) -> None:
    col = po_col(field)
    value = ws.cell(row, col).value
    if not is_blank(value) and not valid_chars(pattern, value):
        proposed = sanitizer(value) if sanitizer and SETTINGS.autofix_invalid_chars else None
        if proposed and valid_chars(pattern, proposed):
            po_add_issue(issues, file, row, col, description, value, proposed)
        else:
            po_add_issue(issues, file, row, col, description, value)


def load_po_reference_lists(workbook) -> dict[str, list[Any]]:
    return {
        "Вид документа": extract_defined_list(workbook, "Вид_документа"),
        "Статус документа": extract_defined_list(workbook, "статус"),
        "Подтверждение утраты": extract_defined_list(workbook, "утратаа"),
        "Подтверждение обмена": extract_defined_list(workbook, "Ообмен"),
        "Подтверждение уничтожения": extract_defined_list(workbook, "унич"),
        "Программа профессионального обучения, направление подготовки": extract_defined_list(workbook, "проф"),
        "Наименование профессий рабочих, должностей служащих": extract_defined_list(workbook, "Квалиф"),
        "Присвоенный квалификационный разряд, класс, категория (при наличии)": extract_defined_list(workbook, "Класс1"),
        "Пол получателя": extract_defined_list(workbook, "пол"),
        "Гражданство получателя (код страны по ОКСМ)": extract_defined_list(workbook, "гражданство"),
        "Форма обучения": extract_defined_list(workbook, "ФО"),
        "Источник финансирования обучения": extract_defined_list(workbook, "финансирование"),
        "Форма получения образования на момент прекращения образовательных отношений": extract_defined_list(workbook, "Отношения"),
        "Наименование документа об образовании (оригинала)": extract_defined_list(workbook, "Вид_документа"),
    }


def scan_po_workbook(path: Path, workbook=None) -> list[Issue]:
    issues: list[Issue] = []
    own_workbook = workbook is None
    if own_workbook:
        workbook = load_workbook(path, read_only=False, data_only=False)

    if SHEET_NAME not in workbook.sheetnames:
        issues.append(Issue(path, 0, 0, "", f"Нет листа '{SHEET_NAME}'", profile="ПО"))
        return issues

    visible_sheets = [ws.title for ws in workbook.worksheets if ws.sheet_state == "visible"]
    if SETTINGS.check_visible_sheet and visible_sheets != [SHEET_NAME]:
        issues.append(Issue(path, 0, 0, "", "В файле должен быть один видимый лист с данными: 'Шаблон'", profile="ПО"))

    ws = workbook[SHEET_NAME]
    refs = load_po_reference_lists(workbook)
    check_sheet_protection(issues, path, ws, "ПО")
    check_column_structure(issues, path, ws, PO_HEADERS, po_row_has_data, "ПО", set(refs))

    headers = [clean_text(ws.cell(1, col).value) for col in range(1, len(PO_HEADERS) + 1)]
    for col, (actual, expected) in enumerate(zip(headers, PO_HEADERS), start=1):
        if normalized_key(actual) != normalized_key(expected):
            po_add_issue(issues, path, 1, col, f"Неверный заголовок, ожидается: {expected}", actual, expected)

    always_required = [
        "Вид документа",
        "Статус документа",
        "Подтверждение утраты",
        "Подтверждение обмена",
        "Подтверждение уничтожения",
        "Дата выдачи документа",
        "Регистрационный номер",
        "Наименование программы профессионального обучения",
        "Наименование профессий рабочих, должностей служащих",
        "Год начала обучения",
        "Год окончания обучения",
        "Срок обучения, часов",
        "Фамилия получателя",
        "Имя получателя",
        "Отчество получателя",
        "Дата рождения получателя",
        "Пол получателя",
    ]

    art_doc = "Свидетельство об освоении дополнительных предпрофессиональных программ в области искусств"
    study_ref = "Справка об обучении"

    for row in range(2, ws.max_row + 1):
        if not po_row_has_data(ws, row):
            continue

        for col in range(1, len(PO_HEADERS) + 1):
            value = ws.cell(row, col).value
            cleaned = normalize_dash(clean_text(value))
            if isinstance(value, str) and value != cleaned:
                po_add_issue(issues, path, row, col, "Лишние пробелы или длинное тире", value, proposed_if_enabled(cleaned, "autofix_whitespace"))

        for field in always_required:
            po_check_required(issues, path, ws, row, field)

        doc_type = clean_text(ws.cell(row, po_col("Вид документа")).value)
        status = clean_text(ws.cell(row, po_col("Статус документа")).value)
        issue_date = parse_date(ws.cell(row, po_col("Дата выдачи документа")).value)

        if normalized_key(doc_type) != normalized_key(study_ref):
            series_col = po_col("Серия документа")
            if is_blank(ws.cell(row, series_col).value):
                po_add_issue(issues, path, row, series_col, "Серия документа не заполнена; при отсутствии серии по инструкции указывается 'Нет'", ws.cell(row, series_col).value, proposed_if_enabled("Нет", "autofix_reference_format"))
            po_check_required(issues, path, ws, row, "Номер документа")

        if normalized_key(doc_type) == normalized_key(art_doc):
            for field in [
                "Программа профессионального обучения, направление подготовки",
                "Наименование профессий рабочих, должностей служащих",
            ]:
                col = po_col(field)
                if not is_blank(ws.cell(row, col).value):
                    po_add_issue(issues, path, row, col, "Для свидетельства по предпрофессиональным программам в области искусств поле должно быть пустым", ws.cell(row, col).value, "")
        elif normalized_key(doc_type) == normalized_key(study_ref):
            field = "Наименование профессий рабочих, должностей служащих"
            col = po_col(field)
            expected = "Дополнительные предпрофессиональные программы в области искусств"
            if is_blank(ws.cell(row, col).value):
                po_add_issue(issues, path, row, col, "Для справки об обучении по программам искусств нужно указать значение из инструкции", ws.cell(row, col).value, expected)

        if issue_date and issue_date.year >= 2021:
            for field in [
                "СНИЛС",
                "Гражданство получателя (код страны по ОКСМ)",
                "Форма обучения",
                "Источник финансирования обучения",
                "Форма получения образования на момент прекращения образовательных отношений",
            ]:
                po_check_required(issues, path, ws, row, field)

        suppressed_cols: set[int] = set()
        if normalized_key(status) == normalized_key("Дубликат"):
            for field in PO_HEADERS[26:34]:
                po_check_required(issues, path, ws, row, field)
        elif normalized_key(status) == normalized_key("Оригинал"):
            for field in PO_HEADERS[26:34]:
                col = po_col(field)
                if not is_blank(ws.cell(row, col).value):
                    po_add_issue(issues, path, row, col, "Сведения об оригинале заполняются только для статуса 'Дубликат'", ws.cell(row, col).value, proposed_if_enabled("", "autofix_conditional_cleanup"))
                    suppressed_cols.add(col)

        change_col = po_col("Номер документа для изменения")
        if not is_blank(ws.cell(row, change_col).value):
            po_add_issue(issues, path, row, change_col, "Поле должно всегда оставаться пустым", ws.cell(row, change_col).value, proposed_if_enabled("", "autofix_conditional_cleanup"))
            suppressed_cols.add(change_col)

        for field, allowed in refs.items():
            col = po_col(field)
            if col in suppressed_cols:
                continue
            if field == "Присвоенный квалификационный разряд, класс, категория (при наличии)":
                is_category_list, split_categories = valid_category_list(ws.cell(row, col).value, allowed)
                if is_category_list:
                    po_add_issue(issues, path, row, col, "Список категорий должен быть разделён на отдельные строки", ws.cell(row, col).value, proposed_if_enabled(split_categories, "autofix_category_split"))
                    continue
            ok, proposed = list_match(ws.cell(row, col).value, allowed)
            if proposed is not None:
                po_add_issue(issues, path, row, col, "Значение отличается от справочника только форматом", ws.cell(row, col).value, proposed_if_enabled(proposed, "autofix_reference_format"))
            elif not ok:
                po_add_issue(issues, path, row, col, "Значение отсутствует в справочнике", ws.cell(row, col).value)

        for field in ["Дата выдачи документа", "Дата рождения получателя", "Дата выдачи (оригинала)"]:
            col = po_col(field)
            if col in suppressed_cols:
                continue
            value = ws.cell(row, col).value
            if is_blank(value):
                continue
            parsed = parse_date(value)
            if not parsed:
                po_add_issue(issues, path, row, col, "Неверный формат даты, нужен дд.мм.гггг", value)
            elif not isinstance(value, (datetime, date)) and clean_text(value) != parsed.strftime("%d.%m.%Y"):
                po_add_issue(issues, path, row, col, "Дата будет приведена к формату дд.мм.гггг", value, proposed_if_enabled(parsed, "autofix_dates_years"))

        if issue_date and not (date(1950, 1, 1) <= issue_date <= date(2100, 12, 31)):
            po_add_issue(issues, path, row, po_col("Дата выдачи документа"), "Дата выдачи вне допустимого диапазона 1950-2100", ws.cell(row, po_col("Дата выдачи документа")).value)

        birth_date = parse_date(ws.cell(row, po_col("Дата рождения получателя")).value)
        if birth_date and not (date(1900, 1, 1) <= birth_date <= date(2100, 12, 31)):
            po_add_issue(issues, path, row, po_col("Дата рождения получателя"), "Дата рождения вне допустимого диапазона 1900-2100", ws.cell(row, po_col("Дата рождения получателя")).value)

        start_year = parse_year(ws.cell(row, po_col("Год начала обучения")).value)
        end_year = parse_year(ws.cell(row, po_col("Год окончания обучения")).value)
        min_start = 1955 if normalized_key(status) == normalized_key("Дубликат") else 1978
        for field, year in [("Год начала обучения", start_year), ("Год окончания обучения", end_year)]:
            col = po_col(field)
            raw = ws.cell(row, col).value
            if is_blank(raw):
                continue
            if year is None:
                po_add_issue(issues, path, row, col, "Нужен год в формате гггг", raw)
            elif text(raw) != str(year):
                po_add_issue(issues, path, row, col, "Год будет приведён к формату гггг", raw, proposed_if_enabled(year, "autofix_dates_years"))
        if start_year:
            if start_year < min_start:
                po_add_issue(issues, path, row, po_col("Год начала обучения"), f"Год начала обучения раньше допустимого минимума {min_start}", start_year)
            if start_year > CURRENT_YEAR:
                po_add_issue(issues, path, row, po_col("Год начала обучения"), f"Год начала обучения больше текущего года {CURRENT_YEAR}", start_year)
        if end_year:
            if end_year > CURRENT_YEAR:
                po_add_issue(issues, path, row, po_col("Год окончания обучения"), f"Год окончания больше текущего года {CURRENT_YEAR}", end_year)
            if start_year and end_year < start_year:
                po_add_issue(issues, path, row, po_col("Год окончания обучения"), "Год окончания меньше года начала обучения", end_year)

        hours_col = po_col("Срок обучения, часов")
        hours_raw = ws.cell(row, hours_col).value
        hours = None
        if isinstance(hours_raw, (int, float)) and int(hours_raw) == hours_raw:
            hours = int(hours_raw)
        elif re.fullmatch(r"\d+", clean_text(hours_raw)):
            hours = int(clean_text(hours_raw))
        if not is_blank(hours_raw):
            if hours is None:
                po_add_issue(issues, path, row, hours_col, "Срок обучения должен быть числом часов", hours_raw)
            elif hours < 6:
                po_add_issue(issues, path, row, hours_col, "Срок обучения для ПО должен быть не менее 6 часов", hours_raw)

        length_limits = {
            "Серия документа": 20,
            "Номер документа": 40,
            "Регистрационный номер": 30,
            "Наименование программы профессионального обучения": 255,
            "Фамилия получателя": 50,
            "Имя получателя": 50,
            "Отчество получателя": 50,
            "Серия (оригинала)": 20,
            "Номер (оригинала)": 20,
            "Регистрационный N (оригинала)": 20,
            "Фамилия получателя (оригинала)": 50,
            "Имя получателя (оригинала)": 50,
            "Отчество получателя (оригинала)": 50,
        }
        for field, limit in length_limits.items():
            if po_col(field) not in suppressed_cols:
                po_check_length(issues, path, ws, row, field, limit)

        name_pattern = r"[А-ЯЁа-яё .'\-()]+"
        for field in ["Фамилия получателя", "Имя получателя", "Отчество получателя", "Фамилия получателя (оригинала)", "Имя получателя (оригинала)", "Отчество получателя (оригинала)"]:
            col = po_col(field)
            if col in suppressed_cols:
                continue
            value = ws.cell(row, col).value
            if not is_blank(value):
                if not valid_chars(name_pattern, value):
                    po_add_issue(issues, path, row, col, "Недопустимые символы: разрешены кириллица, пробел, дефис, точка, апостроф, скобки", value, sanitize_name(value))
                if has_bad_edges_or_repeats(value):
                    po_add_issue(issues, path, row, col, "Пробел/дефис в начале или конце либо повтор пробелов/дефисов", value, clean_text(value).strip("- "))

        for field in ["Серия документа", "Номер документа", "Серия (оригинала)", "Номер (оригинала)"]:
            if po_col(field) not in suppressed_cols:
                po_check_regex(issues, path, ws, row, field, r"[А-ЯЁа-яёA-Z0-9 .\-/]+", "Недопустимые символы: разрешены кириллица, латиница, цифры, точка, дефис, пробел, слэш", sanitize_series)
        for field in ["Регистрационный номер", "Регистрационный N (оригинала)"]:
            if po_col(field) not in suppressed_cols:
                po_check_regex(issues, path, ws, row, field, r"[А-ЯЁа-яёA-Z0-9№()./\-_ ]+", "Недопустимые символы в регистрационном номере", sanitize_po_registration)
        po_check_regex(issues, path, ws, row, "Наименование программы профессионального обучения", r"[А-ЯЁа-яёA-Z0-9№\"«»().,?:/\- &_#+;]+", "Недопустимые символы в наименовании программы", sanitize_po_program)

        snils_col = po_col("СНИЛС")
        snils = clean_text(ws.cell(row, snils_col).value)
        if snils:
            digits = re.sub(r"\D", "", snils)
            if len(digits) == 11 and snils_checksum_ok(digits):
                formatted = format_snils(digits)
                if snils != formatted:
                    po_add_issue(issues, path, row, snils_col, "СНИЛС будет приведён к формату 000-000-000 00", ws.cell(row, snils_col).value, proposed_if_enabled(formatted, "autofix_snils_format"))
            else:
                po_add_issue(issues, path, row, snils_col, "Неверный СНИЛС: нужен формат 000-000-000 00 и корректная контрольная сумма", ws.cell(row, snils_col).value)

    check_formula_columns(issues, path, ws, PO_HEADERS, po_row_has_data, "ПО")

    if own_workbook:
        workbook.close()
    return issues


PROFILES: dict[str, TemplateProfile] = {
    "spo": TemplateProfile("spo", "СПО", EXPECTED_HEADERS, scan_spo_workbook),
    "po": TemplateProfile("po", "ПО", PO_HEADERS, scan_po_workbook),
}


def detect_profile_from_workbook(workbook) -> TemplateProfile:
    if SHEET_NAME not in workbook.sheetnames:
        return PROFILES["spo"]
    ws = workbook[SHEET_NAME]
    first_header = clean_text(ws.cell(1, 1).value)
    column_count = ws.max_column
    if normalized_key(first_header) == normalized_key("Вид документа") and column_count == len(PO_HEADERS):
        return PROFILES["po"]
    return PROFILES["spo"]


def detect_profile(path: Path) -> str:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        return detect_profile_from_workbook(workbook).code
    finally:
        workbook.close()


def deduplicate_issues(issues: list[Issue]) -> list[Issue]:
    result: list[Issue] = []
    seen: set[tuple[Any, ...]] = set()
    for issue in issues:
        key = (
            issue.file,
            issue.row,
            issue.col,
            issue.field,
            value_for_display(issue.current),
            value_for_display(issue.proposed),
            issue.status,
        )
        if key in seen:
            continue
        seen.add(key)
        result.append(issue)
    return result


def scan_workbook(path: Path) -> list[Issue]:
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        profile = detect_profile_from_workbook(workbook)
        return deduplicate_issues(profile.scanner(path, workbook))
    finally:
        workbook.close()


def backup_path_for(path: Path) -> Path:
    backup_dir = path.parent / "_bak"
    backup_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return backup_dir / f"{path.stem}_{timestamp}{path.suffix}.bak"


def make_backup(path: Path) -> Path:
    backup = backup_path_for(path)
    shutil.copy2(path, backup)
    return backup


def copy_row(ws, src_row: int, dst_row: int) -> None:
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for col in range(1, ws.max_column + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        dst.value = src.value
        if src.has_style:
            dst._style = copy.copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.font:
            dst.font = copy.copy(src.font)
        if src.fill:
            dst.fill = copy.copy(src.fill)
        if src.border:
            dst.border = copy.copy(src.border)
        if src.alignment:
            dst.alignment = copy.copy(src.alignment)
        if src.protection:
            dst.protection = copy.copy(src.protection)
        if src.comment:
            dst.comment = copy.copy(src.comment)


def category_split_groups(issues: list[Issue]) -> list[tuple[int, int, tuple[str, ...]]]:
    split_issues = sorted(
        (issue for issue in issues if isinstance(issue.proposed, CategorySplit)),
        key=lambda issue: issue.row,
    )
    groups: list[tuple[int, int, tuple[str, ...]]] = []
    for issue in split_issues:
        parts = issue.proposed.parts
        if groups and groups[-1][1] + 1 == issue.row and groups[-1][2] == parts:
            start, _end, old_parts = groups[-1]
            groups[-1] = (start, issue.row, old_parts)
        else:
            groups.append((issue.row, issue.row, parts))
    return groups


def apply_category_splits(ws, issues: list[Issue]) -> None:
    category_col = po_col("Присвоенный квалификационный разряд, класс, категория (при наличии)")
    for start_row, end_row, parts in reversed(category_split_groups(issues)):
        row_count = end_row - start_row + 1
        for row in range(start_row, end_row + 1):
            ws.cell(row, category_col).value = parts[0]
        for category in reversed(parts[1:]):
            insert_at = end_row + 1
            ws.insert_rows(insert_at, row_count)
            for offset in range(row_count):
                src_row = start_row + offset
                dst_row = insert_at + offset
                copy_row(ws, src_row, dst_row)
                ws.cell(dst_row, category_col).value = category


def apply_fixes(issues: list[Issue]) -> dict[Path, int]:
    by_file: dict[Path, list[Issue]] = {}
    for issue in issues:
        if issue.proposed is not None and ((issue.row and issue.col) or issue.action == "protect_sheet"):
            by_file.setdefault(issue.file, []).append(issue)

    saved: dict[Path, int] = {}
    for path, file_issues in by_file.items():
        make_backup(path)
        workbook = load_workbook(path, read_only=False, data_only=False)
        ws = workbook[SHEET_NAME]
        normal_issues = [issue for issue in file_issues if issue.action == "cell" and not isinstance(issue.proposed, CategorySplit)]
        split_issues = [issue for issue in file_issues if isinstance(issue.proposed, CategorySplit)]
        protect_issues = [issue for issue in file_issues if issue.action == "protect_sheet"]
        unhide_issues = [issue for issue in file_issues if issue.action == "unhide_column"]
        for issue in normal_issues:
            cell = ws.cell(issue.row, issue.col)
            cell.value = issue.proposed
            if isinstance(issue.proposed, (date, datetime)):
                cell.number_format = "DD.MM.YYYY"
            elif is_formula(issue.proposed):
                cell.number_format = "General"
            elif issue.field in ("Год поступления", "Год окончания", "Срок обучения, лет", "Год начала обучения", "Год окончания обучения", "Срок обучения, часов"):
                cell.number_format = "0"
        if protect_issues:
            ws.protection.sheet = True
        for issue in unhide_issues:
            ws.column_dimensions[get_column_letter(issue.col)].hidden = False
        if split_issues:
            apply_category_splits(ws, split_issues)
        workbook.save(path)
        workbook.close()
        saved[path] = len(file_issues)
    return saved


def reference_values_for_issue(issue: Issue) -> list[str]:
    if not issue.row or not issue.col or not issue.file.exists():
        return []
    workbook = load_workbook(issue.file, read_only=False, data_only=False)
    try:
        profile = detect_profile_from_workbook(workbook)
        if profile.code == "po":
            refs = load_po_reference_lists(workbook)
        else:
            refs = load_reference_lists(workbook)
        values = refs.get(issue.field, [])
        return [value_for_display(value) for value in values if not is_blank(value)]
    finally:
        workbook.close()


def manual_field_kind(issue: Issue) -> str:
    field = issue.field.casefold()
    message = issue.message.casefold()
    if "снилс" in field or "снилс" in message:
        return "snils"
    if "дата" in field or "дата" in message:
        return "date"
    if "год" in field or "год" in message:
        return "year"
    if "час" in field or "час" in message:
        return "hours"
    if "только цифры" in message or "огрн" in field.casefold() or "кпп" in field.casefold():
        return "digits"
    return "text"


def validate_manual_value(issue: Issue, raw_value: str) -> tuple[bool, str, Any]:
    value = raw_value.strip()
    if value == "":
        return True, "", None
    kind = manual_field_kind(issue)
    if kind == "date":
        parsed = parse_date(value)
        if not parsed:
            return False, "Введите дату в формате дд.мм.гггг.", value
        return True, "", parsed
    if kind == "year":
        year = parse_year(value)
        if year is None:
            return False, "Введите год четырьмя цифрами.", value
        return True, "", year
    if kind == "hours":
        if not re.fullmatch(r"\d+", value):
            return False, "Введите количество часов целым числом.", value
        return True, "", int(value)
    if kind == "snils":
        digits = re.sub(r"\D", "", value)
        if len(digits) != 11 or not snils_checksum_ok(digits):
            return False, "СНИЛС должен содержать 11 цифр и корректную контрольную сумму.", value
        return True, "", format_snils(digits)
    if kind == "digits" and not re.fullmatch(r"\d+", value):
        return False, "Поле должно содержать только цифры.", value
    return True, "", value


def manual_editor_hint(issue: Issue) -> str:
    kind = manual_field_kind(issue)
    return {
        "date": "Формат: дд.мм.гггг",
        "year": "Формат: гггг",
        "hours": "Целое число часов",
        "snils": "Формат: 000-000-000 00; контрольная сумма проверяется до сохранения",
        "digits": "Только цифры",
    }.get(kind, "Введите значение или выберите вариант из справочника.")


def apply_manual_edit(issue: Issue, new_value: Any) -> Path:
    if not issue.row or not issue.col:
        raise ValueError("У выбранной проблемы нет конкретной ячейки.")
    backup = make_backup(issue.file)
    workbook = load_workbook(issue.file, read_only=False, data_only=False)
    try:
        ws = workbook[SHEET_NAME]
        cell = ws.cell(issue.row, issue.col)
        cell.value = new_value
        if isinstance(new_value, (date, datetime)):
            cell.number_format = "DD.MM.YYYY"
        elif is_formula(new_value):
            cell.number_format = "General"
        elif isinstance(new_value, int) and manual_field_kind(issue) in {"year", "hours"}:
            cell.number_format = "0"
        workbook.save(issue.file)
    finally:
        workbook.close()
    return backup


def fix_report_rows(by_file: dict[Path, list[Issue]], relative_name: Callable[[Path], str] | None = None) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for path, file_issues in sorted(by_file.items(), key=lambda item: str(item[0]).casefold()):
        kinds = Counter(issue.message for issue in file_issues)
        rows.append(
            {
                "file": relative_name(path) if relative_name else path.name,
                "count": len(file_issues),
                "details": kinds.most_common(),
            }
        )
    return rows


def save_fix_report_doc(report_path: Path, by_file: dict[Path, list[Issue]], relative_name: Callable[[Path], str] | None = None) -> None:
    rows = fix_report_rows(by_file, relative_name)
    total = sum(row["count"] for row in rows)
    generated_at = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
    body_rows = []
    for index, row in enumerate(rows, start=1):
        details = "<br>".join(
            f"{html.escape(message)}: {count}"
            for message, count in row["details"]
        )
        body_rows.append(
            "<tr>"
            f"<td>{index}</td>"
            f"<td>{html.escape(row['file'])}</td>"
            f"<td>{row['count']}</td>"
            f"<td>{details}</td>"
            "</tr>"
        )
    document = f"""<!doctype html>
<html>
<head>
<meta charset="utf-8">
<title>Отчёт по автоматическим исправлениям ФРДО</title>
<style>
body {{ font-family: Calibri, Arial, sans-serif; font-size: 11pt; color: #1f2937; }}
h1 {{ font-size: 18pt; margin-bottom: 4pt; }}
.meta {{ color: #4b5563; margin-bottom: 14pt; }}
table {{ border-collapse: collapse; width: 100%; }}
th, td {{ border: 1px solid #9ca3af; padding: 6px; vertical-align: top; }}
th {{ background: #e5e7eb; }}
td:nth-child(1), td:nth-child(3) {{ text-align: center; }}
</style>
</head>
<body>
<h1>Отчёт по автоматическим исправлениям ФРДО</h1>
<div class="meta">Дата формирования: {generated_at}<br>Файлов в отчёте: {len(rows)}<br>Всего исправлений: {total}</div>
<table>
<thead>
<tr><th>№</th><th>Файл</th><th>Исправлений</th><th>Типы исправлений</th></tr>
</thead>
<tbody>
{''.join(body_rows)}
</tbody>
</table>
</body>
</html>
"""
    report_path.write_text(document, encoding="utf-8")


def save_issue_table_report_doc(report_path: Path, issues: list[Issue], relative_name: Callable[[Path], str] | None = None) -> None:
    generated_at = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
    body_rows = []
    for index, issue in enumerate(issues, start=1):
        file_name = relative_name(issue.file) if relative_name else issue.file.name
        body_rows.append(
            "<tr>"
            f"<td>{index}</td>"
            f"<td>{html.escape(issue.profile)}</td>"
            f"<td>{html.escape(issue.severity)}</td>"
            f"<td>{html.escape(issue.status)}</td>"
            f"<td>{html.escape(file_name)}</td>"
            f"<td>{html.escape(issue.address)}</td>"
            f"<td>{html.escape(issue.field)}</td>"
            f"<td>{html.escape(issue.message)}</td>"
            f"<td>{html.escape(value_for_display(issue.current))}</td>"
            "</tr>"
        )
    document = f"""<!doctype html>
<html>
<head>
<meta charset="utf-8">
<title>Отчёт по таблице ошибок ФРДО</title>
<style>
@page Section1 {{
    size: 29.7cm 21cm;
    mso-page-orientation: landscape;
    margin: 0.7cm 0.7cm 0.7cm 0.7cm;
}}
div.Section1 {{ page: Section1; }}
body {{ font-family: Calibri, Arial, sans-serif; font-size: 9pt; color: #1f2937; }}
h1 {{ font-size: 18pt; margin-bottom: 4pt; }}
.meta {{ color: #4b5563; margin-bottom: 14pt; }}
table {{
    border-collapse: collapse;
    width: 100%;
    table-layout: fixed;
    font-size: 8pt;
}}
thead {{ display: table-header-group; }}
tr.header-row {{ mso-yfti-irow: 0; mso-yfti-firstrow: yes; mso-table-header-row: yes; }}
th, td {{
    border: 1px solid #9ca3af;
    padding: 3px;
    vertical-align: top;
    overflow-wrap: anywhere;
    word-wrap: break-word;
    word-break: break-word;
}}
th {{ background: #e5e7eb; font-weight: bold; }}
.c-num {{ width: 3%; }}
.c-profile {{ width: 5%; }}
.c-severity {{ width: 8%; }}
.c-status {{ width: 7%; }}
.c-file {{ width: 15%; }}
.c-cell {{ width: 5%; }}
.c-field {{ width: 18%; }}
.c-message {{ width: 29%; }}
.c-current {{ width: 10%; }}
</style>
</head>
<body>
<div class="Section1">
<h1>Отчёт по таблице ошибок ФРДО</h1>
<div class="meta">Дата формирования: {generated_at}<br>Строк в отчёте: {len(issues)}</div>
<table>
<colgroup>
<col class="c-num"><col class="c-profile"><col class="c-severity"><col class="c-status"><col class="c-file"><col class="c-cell"><col class="c-field"><col class="c-message"><col class="c-current">
</colgroup>
<thead>
<tr>
<th>№</th><th>Профиль</th><th>Важность</th><th>Статус</th><th>Файл</th><th>Ячейка</th><th>Поле</th><th>Проблема</th><th>Сейчас</th>
</tr>
</thead>
<tbody>
{''.join(body_rows)}
</tbody>
</table>
</div>
</body>
</html>
"""
    report_path.write_text(document, encoding="utf-8")


def duplicate_value(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return normalized_key(value)


def register_duplicate(registry: dict[tuple[str, str, tuple[str, ...]], list[Issue]], key: tuple[str, str, tuple[str, ...]], issue: Issue) -> None:
    if all(part for part in key[2]):
        registry.setdefault(key, []).append(issue)


def duplicate_entries_for_file(path: Path) -> list[tuple[tuple[str, str, tuple[str, ...]], Issue]]:
    entries: list[tuple[tuple[str, str, tuple[str, ...]], Issue]] = []
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        profile = detect_profile_from_workbook(workbook)
        if SHEET_NAME not in workbook.sheetnames:
            return entries
        ws = workbook[SHEET_NAME]
        if profile.code == "po":
            for row in range(2, ws.max_row + 1):
                if not po_row_has_data(ws, row):
                    continue
                category = duplicate_value(ws.cell(row, po_col("Присвоенный квалификационный разряд, класс, категория (при наличии)")).value)
                series = duplicate_value(ws.cell(row, po_col("Серия документа")).value)
                number = duplicate_value(ws.cell(row, po_col("Номер документа")).value)
                snils = re.sub(r"\D", "", text(ws.cell(row, po_col("СНИЛС")).value))
                person = tuple(
                    duplicate_value(ws.cell(row, po_col(field)).value)
                    for field in ["Фамилия получателя", "Имя получателя", "Отчество получателя", "Дата рождения получателя"]
                )
                entries.append((("ПО", "документ", (series, number, category)), Issue(path, row, po_col("Номер документа"), "Номер документа", "Дублирующийся номер документа в той же категории", ws.cell(row, po_col("Номер документа")).value, profile="ПО")))
                entries.append((("ПО", "СНИЛС", (snils, category)), Issue(path, row, po_col("СНИЛС"), "СНИЛС", "Дублирующийся СНИЛС в той же категории", ws.cell(row, po_col("СНИЛС")).value, profile="ПО")))
                entries.append((("ПО", "получатель", (*person, category)), Issue(path, row, po_col("Фамилия получателя"), "ФИО + дата рождения", "Дублирующийся получатель в той же категории", " ".join(value_for_display(ws.cell(row, po_col(field)).value) for field in ["Фамилия получателя", "Имя получателя", "Отчество получателя", "Дата рождения получателя"]), profile="ПО")))
        else:
            for row in range(2, ws.max_row + 1):
                if not row_has_data(ws, row):
                    continue
                series = duplicate_value(ws.cell(row, COL["Серия документа"]).value)
                number = duplicate_value(ws.cell(row, COL["Номер документа"]).value)
                snils = re.sub(r"\D", "", text(ws.cell(row, COL["СНИЛС"]).value))
                person = tuple(
                    duplicate_value(ws.cell(row, COL[field]).value)
                    for field in ["Фамилия получателя", "Имя получателя", "Отчество получателя", "Дата рождения получателя"]
                )
                entries.append((("СПО", "документ", (series, number)), Issue(path, row, COL["Номер документа"], "Номер документа", "Дублирующийся номер документа", ws.cell(row, COL["Номер документа"]).value, profile="СПО")))
                entries.append((("СПО", "СНИЛС", (snils,)), Issue(path, row, COL["СНИЛС"], "СНИЛС", "Дублирующийся СНИЛС", ws.cell(row, COL["СНИЛС"]).value, profile="СПО")))
                entries.append((("СПО", "получатель", (*person,)), Issue(path, row, COL["Фамилия получателя"], "ФИО + дата рождения", "Дублирующийся получатель", " ".join(value_for_display(ws.cell(row, COL[field]).value) for field in ["Фамилия получателя", "Имя получателя", "Отчество получателя", "Дата рождения получателя"]), profile="СПО")))
    finally:
        workbook.close()
    return entries


def worker_count(file_count: int) -> int:
    if not SETTINGS.parallel_scan or file_count < 2:
        return 1
    return max(1, min(8, file_count, (os.cpu_count() or 4)))


def duplicate_issues_for_files(files: list[Path]) -> list[Issue]:
    registry: dict[tuple[str, str, tuple[str, ...]], list[Issue]] = {}
    workers = worker_count(len(files))
    if workers == 1:
        file_entries = [duplicate_entries_for_file(path) for path in files]
    else:
        file_entries = []
        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = [executor.submit(duplicate_entries_for_file, path) for path in files]
            for future in as_completed(futures):
                file_entries.append(future.result())

    for entries in file_entries:
        for key, issue in entries:
            register_duplicate(registry, key, issue)

    duplicates: list[Issue] = []
    for occurrences in registry.values():
        if len(occurrences) > 1:
            duplicates.extend(occurrences)
    return duplicates


def scan_files(files: list[Path]) -> list[Issue]:
    issues: list[Issue] = []
    workers = worker_count(len(files))
    if workers == 1:
        for path in files:
            try:
                issues.extend(scan_workbook(path))
            except Exception as exc:
                issues.append(Issue(path, 0, 0, "", f"Не удалось проверить файл: {exc}"))
                traceback.print_exc()
        return issues

    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {executor.submit(scan_workbook, path): path for path in files}
        for future in as_completed(futures):
            path = futures[future]
            try:
                issues.extend(future.result())
            except Exception as exc:
                issues.append(Issue(path, 0, 0, "", f"Не удалось проверить файл: {exc}"))
                traceback.print_exc()
    return issues


def ps_quote(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def open_issue_in_excel(issue: Issue) -> None:
    if not issue.file.exists():
        raise FileNotFoundError(issue.file)
    if not issue.address or sys.platform != "win32":
        os.startfile(issue.file)
        return
    path = str(issue.file.resolve())
    script = f"""
$path = {ps_quote(path)}
$sheet = {ps_quote(SHEET_NAME)}
$addr = {ps_quote(issue.address)}
try {{
    $excel = [Runtime.InteropServices.Marshal]::GetActiveObject('Excel.Application')
}} catch {{
    $excel = New-Object -ComObject Excel.Application
}}
$excel.Visible = $true
$wb = $null
foreach ($book in @($excel.Workbooks)) {{
    if ($book.FullName -ieq $path) {{ $wb = $book; break }}
}}
if ($null -eq $wb) {{ $wb = $excel.Workbooks.Open($path) }}
$ws = $wb.Worksheets.Item($sheet)
$ws.Activate()
$range = $ws.Range($addr)
$range.Select()
$excel.ActiveWindow.ScrollRow = [Math]::Max(1, $range.Row - 5)
$excel.ActiveWindow.ScrollColumn = [Math]::Max(1, $range.Column - 3)
"""
    creationflags = subprocess.CREATE_NO_WINDOW if hasattr(subprocess, "CREATE_NO_WINDOW") else 0
    subprocess.Popen(
        ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script],
        creationflags=creationflags,
    )


class CheckerApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title(APP_TITLE)
        self.geometry_for_screen()
        self.minsize(1280, 720)
        self.directory = Path.cwd()
        self.issues: list[Issue] = []
        self.filtered_issues: list[Issue] = []
        self.sort_column: str | None = None
        self.sort_reverse = False
        self.status_var = tk.StringVar(value="Готово")
        self._configure_style()
        self._build_ui()
        self.scan()

    def geometry_for_screen(self) -> None:
        screen_w = self.winfo_screenwidth()
        screen_h = self.winfo_screenheight()
        width = min(screen_w, max(1280, int(screen_w * 0.98)))
        height = min(screen_h - 40, max(720, int(screen_h * 0.94)))
        x = max(0, (screen_w - width) // 2)
        y = max(0, (screen_h - height) // 2)
        self.geometry(f"{width}x{height}+{x}+{y}")
        if sys.platform == "win32" and screen_w >= 1600 and screen_h >= 900:
            try:
                self.state("zoomed")
            except tk.TclError:
                pass

    def _configure_style(self) -> None:
        self.configure(background="#f5f7fb")
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        style.configure(".", font=("Segoe UI", 10), background="#f5f7fb", foreground="#172033")
        style.configure("TFrame", background="#f5f7fb")
        style.configure("Toolbar.TFrame", background="#eef2f7")
        style.configure("TLabel", background="#f5f7fb", foreground="#172033")
        style.configure("Muted.TLabel", background="#f5f7fb", foreground="#5d667a")
        style.configure("TButton", padding=(10, 6), background="#ffffff", foreground="#172033", borderwidth=1)
        style.map("TButton", background=[("active", "#e9eef7")])
        style.configure("Accent.TButton", background="#2563eb", foreground="#ffffff")
        style.map("Accent.TButton", background=[("active", "#1d4ed8")], foreground=[("active", "#ffffff")])
        style.configure("TCombobox", padding=4)
        style.configure("Treeview", rowheight=28, fieldbackground="#ffffff", background="#ffffff", foreground="#172033", borderwidth=0)
        style.configure("Treeview.Heading", padding=(8, 7), background="#e6ebf3", foreground="#172033", font=("Segoe UI", 10, "bold"))
        style.map("Treeview", background=[("selected", "#cfe0ff")], foreground=[("selected", "#111827")])

    def _build_ui(self) -> None:
        top = ttk.Frame(self, padding=(10, 8), style="Toolbar.TFrame")
        top.pack(fill=tk.X)

        self.dir_var = tk.StringVar(value=str(self.directory))
        ttk.Label(top, text="Папка:").pack(side=tk.LEFT)
        ttk.Entry(top, textvariable=self.dir_var).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=8)
        ttk.Button(top, text="Выбрать", command=self.choose_dir).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(top, text="Проверить", command=self.scan, style="Accent.TButton").pack(side=tk.LEFT, padx=(0, 6))

        actions = ttk.Frame(self, padding=(10, 8, 10, 6), style="Toolbar.TFrame")
        actions.pack(fill=tk.X)
        ttk.Button(actions, text="Исправить выбранные", command=self.fix_selected).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(actions, text="Исправить в фильтре", command=self.fix_all).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(actions, text="Открыть ячейку", command=self.open_selected_cell).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(actions, text="Править ячейку", command=self.edit_selected_cell).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(actions, text="Отчёт таблицы .doc", command=self.save_table_report).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(actions, text="Настройки правил", command=self.open_settings).pack(side=tk.LEFT, padx=(0, 6))

        filters = ttk.Frame(self, padding=(10, 0, 10, 8))
        filters.pack(fill=tk.X)
        self.status_filter = tk.StringVar(value="Все")
        self.profile_filter = tk.StringVar(value="Все")
        self.severity_filter = tk.StringVar(value="Все")
        self.file_filter = tk.StringVar(value="Все")

        ttk.Label(filters, text="Проблемы:").pack(side=tk.LEFT)
        self.status_combo = ttk.Combobox(filters, textvariable=self.status_filter, state="readonly", width=18, values=("Все", "Только ошибки", "Только исправимые"))
        self.status_combo.pack(side=tk.LEFT, padx=(4, 12))
        ttk.Label(filters, text="Профиль:").pack(side=tk.LEFT)
        self.profile_combo = ttk.Combobox(filters, textvariable=self.profile_filter, state="readonly", width=8, values=("Все", "СПО", "ПО"))
        self.profile_combo.pack(side=tk.LEFT, padx=(4, 12))
        ttk.Label(filters, text="Важность:").pack(side=tk.LEFT)
        self.severity_combo = ttk.Combobox(filters, textvariable=self.severity_filter, state="readonly", width=18, values=("Все", SEVERITY_BLOCKING, SEVERITY_WARNING, SEVERITY_AUTOFIX))
        self.severity_combo.pack(side=tk.LEFT, padx=(4, 12))
        ttk.Label(filters, text="Файл:").pack(side=tk.LEFT)
        self.file_combo = ttk.Combobox(filters, textvariable=self.file_filter, state="readonly", width=34, values=("Все",))
        self.file_combo.pack(side=tk.LEFT, padx=(4, 12))
        ttk.Button(filters, text="Сбросить", command=self.reset_filters).pack(side=tk.LEFT)
        for combo in (self.status_combo, self.profile_combo, self.severity_combo, self.file_combo):
            combo.bind("<<ComboboxSelected>>", lambda _event: self.apply_filters())

        summary = ttk.Frame(self, padding=(10, 0, 10, 8))
        summary.pack(fill=tk.X)
        self.summary_var = tk.StringVar()
        ttk.Label(summary, textvariable=self.summary_var).pack(side=tk.LEFT)

        columns = ("profile", "severity", "status", "file", "cell", "field", "message", "hint", "current", "proposed")
        table_frame = ttk.Frame(self, padding=(10, 0, 10, 6))
        table_frame.pack(fill=tk.BOTH, expand=True)
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", selectmode="extended")
        headings = {
            "profile": "Профиль",
            "severity": "Важность",
            "status": "Статус",
            "file": "Файл",
            "cell": "Ячейка",
            "field": "Поле",
            "message": "Проблема",
            "hint": "Подсказка",
            "current": "Сейчас",
            "proposed": "Исправить на",
        }
        widths = {
            "profile": 64,
            "severity": 122,
            "status": 122,
            "file": 250,
            "cell": 66,
            "field": 270,
            "message": 560,
            "hint": 560,
            "current": 190,
            "proposed": 190,
        }
        for col in columns:
            self.tree.heading(col, text=headings[col], command=lambda c=col: self.sort_by_column(c))
            self.tree.column(col, width=widths[col], minwidth=50, stretch=False)
        self.tree.tag_configure(SEVERITY_BLOCKING, background="#ffe5e5")
        self.tree.tag_configure(SEVERITY_WARNING, background="#fff6cf")
        self.tree.tag_configure(SEVERITY_AUTOFIX, background="#e8f7ea")
        self.tree.grid(row=0, column=0, sticky="nsew")

        yscroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        xscroll = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)
        self.tree.bind("<Double-1>", lambda _event: self.open_selected_cell())
        self.tree.bind("<<TreeviewSelect>>", lambda _event: self.update_detail_panel())

        detail = ttk.Frame(self, padding=(10, 0, 10, 8))
        detail.pack(fill=tk.X)
        self.detail_var = tk.StringVar(value="Выберите строку, чтобы увидеть полный текст ошибки и подсказку.")
        self.detail_label = ttk.Label(detail, textvariable=self.detail_var, wraplength=1800, justify=tk.LEFT)
        self.detail_label.pack(fill=tk.X)
        self.bind("<Configure>", self.on_resize)

        bottom = ttk.Frame(self, padding=(10, 0, 10, 10))
        bottom.pack(fill=tk.X)
        ttk.Label(
            bottom,
            text="Перед сохранением создаётся .bak-копия в папке _bak. Неисправимые ошибки требуют ручной правки в Excel.",
            style="Muted.TLabel",
        ).pack(side=tk.LEFT)
        ttk.Label(bottom, textvariable=self.status_var, style="Muted.TLabel").pack(side=tk.RIGHT)

    def choose_dir(self) -> None:
        selected = filedialog.askdirectory(initialdir=str(self.directory))
        if selected:
            self.directory = Path(selected)
            self.dir_var.set(str(self.directory))
            self.scan()

    def set_busy(self, busy: bool) -> None:
        self.configure(cursor="watch" if busy else "")
        self.status_var.set("Выполняется проверка..." if busy else "Готово")
        self.update_idletasks()

    def on_resize(self, _event: tk.Event) -> None:
        if hasattr(self, "detail_label"):
            self.detail_label.configure(wraplength=max(700, self.winfo_width() - 40))

    def relative_file(self, path: Path) -> str:
        try:
            return str(path.relative_to(self.directory))
        except ValueError:
            return path.name

    def reset_filters(self) -> None:
        self.status_filter.set("Все")
        self.profile_filter.set("Все")
        self.severity_filter.set("Все")
        self.file_filter.set("Все")
        self.apply_filters()

    def refresh_file_filter(self) -> None:
        current = self.file_filter.get()
        files = sorted({self.relative_file(issue.file) for issue in self.issues})
        values = ("Все", *files)
        self.file_combo.configure(values=values)
        self.file_filter.set(current if current in values else "Все")

    def issue_matches_filters(self, issue: Issue) -> bool:
        status = self.status_filter.get()
        if status == "Только ошибки" and issue.proposed is not None:
            return False
        if status == "Только исправимые" and issue.proposed is None:
            return False
        profile = self.profile_filter.get()
        if profile != "Все" and issue.profile != profile:
            return False
        severity = self.severity_filter.get()
        if severity != "Все" and issue.severity != severity:
            return False
        file_name = self.file_filter.get()
        if file_name != "Все" and self.relative_file(issue.file) != file_name:
            return False
        return True

    def sort_key(self, issue: Issue, column: str) -> Any:
        if column == "profile":
            return issue.profile
        if column == "severity":
            order = {SEVERITY_BLOCKING: 0, SEVERITY_WARNING: 1, SEVERITY_AUTOFIX: 2}
            return order.get(issue.severity, 9)
        if column == "status":
            return issue.status
        if column == "file":
            return self.relative_file(issue.file).casefold()
        if column == "cell":
            return (issue.row, issue.col)
        if column == "field":
            return issue.field.casefold()
        if column == "message":
            return issue.message.casefold()
        if column == "hint":
            return issue.hint.casefold()
        if column == "current":
            return value_for_display(issue.current).casefold()
        if column == "proposed":
            return value_for_display(issue.proposed).casefold()
        return ""

    def sort_by_column(self, column: str) -> None:
        if self.sort_column == column:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_column = column
            self.sort_reverse = False
        self.apply_filters()

    def apply_filters(self) -> None:
        self.tree.delete(*self.tree.get_children())
        self.filtered_issues = [issue for issue in self.issues if self.issue_matches_filters(issue)]
        if self.sort_column:
            self.filtered_issues.sort(key=lambda issue: self.sort_key(issue, self.sort_column or ""), reverse=self.sort_reverse)
        for idx, issue in enumerate(self.filtered_issues):
            self.tree.insert(
                "",
                tk.END,
                iid=str(idx),
                tags=(issue.severity,),
                values=(
                    issue.profile,
                    issue.severity,
                    issue.status,
                    self.relative_file(issue.file),
                    issue.address,
                    issue.field,
                    issue.message,
                    issue.hint,
                    value_for_display(issue.current),
                    value_for_display(issue.proposed),
                ),
            )
        total_fixable = sum(1 for issue in self.issues if issue.proposed is not None)
        filtered_fixable = sum(1 for issue in self.filtered_issues if issue.proposed is not None)
        blocking = sum(1 for issue in self.filtered_issues if issue.severity == SEVERITY_BLOCKING)
        self.summary_var.set(
            f"Показано: {len(self.filtered_issues)} из {len(self.issues)}. "
            f"Блокирующих: {blocking}. Исправимых в фильтре: {filtered_fixable}. Всего исправимых: {total_fixable}."
        )
        self.update_detail_panel()

    def scan(self) -> None:
        self.directory = Path(self.dir_var.get()).expanduser()
        self.tree.delete(*self.tree.get_children())
        self.issues = []
        self.filtered_issues = []
        if not self.directory.exists():
            messagebox.showerror(APP_TITLE, "Папка не найдена")
            return
        self.set_busy(True)
        try:
            files = sorted(self.directory.rglob("*.xlsx"))
            files = [path for path in files if not path.name.startswith("~$")]
            self.issues.extend(scan_files(files))
            if SETTINGS.check_duplicates:
                self.issues.extend(duplicate_issues_for_files(files))
            self.issues = deduplicate_issues(self.issues)
            self.refresh_file_filter()
            self.apply_filters()
        finally:
            self.set_busy(False)

    def selected_issues(self) -> list[Issue]:
        return [self.filtered_issues[int(item)] for item in self.tree.selection()]

    def update_detail_panel(self) -> None:
        if not hasattr(self, "detail_var"):
            return
        selected = self.selected_issues()
        if not selected:
            self.detail_var.set("Выберите строку, чтобы увидеть полный текст ошибки и подсказку.")
            return
        issue = selected[0]
        parts = [
            f"{self.relative_file(issue.file)}  {issue.address or ''}".strip(),
            f"Поле: {issue.field or '-'}",
            f"Проблема: {issue.message}",
            f"Подсказка: {issue.hint}",
            f"Сейчас: {value_for_display(issue.current) or '-'}",
            f"Исправить на: {value_for_display(issue.proposed) or '-'}",
        ]
        self.detail_var.set("    |    ".join(parts))

    def fix_selected(self) -> None:
        self.fix(self.selected_issues(), selected=True)

    def fix_all(self) -> None:
        self.fix(self.filtered_issues, selected=False)

    def open_selected_cell(self) -> None:
        selected = self.selected_issues()
        if not selected:
            messagebox.showinfo(APP_TITLE, "Выберите строку с ошибкой.")
            return
        issue = selected[0]
        try:
            open_issue_in_excel(issue)
        except Exception as exc:
            messagebox.showerror(APP_TITLE, f"Не удалось открыть файл: {exc}")
            traceback.print_exc()

    def edit_selected_cell(self) -> None:
        selected = self.selected_issues()
        if not selected:
            messagebox.showinfo(APP_TITLE, "Выберите строку с конкретной ячейкой.")
            return
        issue = selected[0]
        if not issue.row or not issue.col:
            messagebox.showinfo(APP_TITLE, "У выбранной проблемы нет конкретной ячейки для правки.")
            return

        try:
            ref_values = reference_values_for_issue(issue)
        except Exception:
            ref_values = []

        dialog = tk.Toplevel(self)
        dialog.title("Правка ячейки")
        dialog.transient(self)
        dialog.grab_set()
        dialog.geometry("680x330")
        dialog.minsize(560, 300)

        body = ttk.Frame(dialog, padding=14)
        body.pack(fill=tk.BOTH, expand=True)
        ttk.Label(body, text=f"{self.relative_file(issue.file)} / {issue.address}", font=("Segoe UI", 10, "bold")).pack(anchor="w")
        ttk.Label(body, text=issue.field, style="Muted.TLabel").pack(anchor="w", pady=(2, 10))
        ttk.Label(body, text=issue.message, wraplength=630).pack(anchor="w")
        ttk.Label(body, text=issue.hint, style="Muted.TLabel", wraplength=630).pack(anchor="w", pady=(4, 12))

        value_var = tk.StringVar(value=value_for_display(issue.proposed if issue.proposed is not None else issue.current))
        ttk.Label(body, text="Новое значение").pack(anchor="w")
        if ref_values:
            editor = ttk.Combobox(body, textvariable=value_var, values=ref_values, state="normal")
        else:
            editor = ttk.Entry(body, textvariable=value_var)
        editor.pack(fill=tk.X, pady=(4, 8))
        editor.focus_set()

        ttk.Label(body, text=manual_editor_hint(issue), style="Muted.TLabel", wraplength=630).pack(anchor="w")
        ttk.Label(body, text=f"Сейчас: {value_for_display(issue.current)}", style="Muted.TLabel", wraplength=630).pack(anchor="w")

        buttons = ttk.Frame(dialog, padding=(14, 0, 14, 14))
        buttons.pack(fill=tk.X)

        def save() -> None:
            try:
                ok, error, normalized_value = validate_manual_value(issue, value_var.get())
                if not ok:
                    messagebox.showwarning(APP_TITLE, error)
                    return
                backup = apply_manual_edit(issue, normalized_value)
                dialog.destroy()
                self.scan()
                messagebox.showinfo(APP_TITLE, f"Ячейка сохранена. Резервная копия:\n{backup}")
            except Exception as exc:
                messagebox.showerror(APP_TITLE, f"Не удалось сохранить правку: {exc}")
                traceback.print_exc()

        ttk.Button(buttons, text="Отмена", command=dialog.destroy).pack(side=tk.RIGHT)
        ttk.Button(buttons, text="Сохранить", command=save, style="Accent.TButton").pack(side=tk.RIGHT, padx=(0, 6))
        self.wait_window(dialog)

    def save_table_report(self) -> None:
        if not self.filtered_issues:
            messagebox.showinfo(APP_TITLE, "В текущей таблице нет строк для отчёта.")
            return
        default_name = f"frdo_table_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.doc"
        target = filedialog.asksaveasfilename(
            parent=self,
            title="Сохранить отчёт по таблице",
            defaultextension=".doc",
            initialdir=str(self.directory),
            initialfile=default_name,
            filetypes=(("Документ Word", "*.doc"), ("Все файлы", "*.*")),
        )
        if not target:
            return
        try:
            save_issue_table_report_doc(Path(target), self.filtered_issues, self.relative_file)
            messagebox.showinfo(APP_TITLE, f"Отчёт сохранён:\n{target}")
        except Exception as exc:
            messagebox.showerror(APP_TITLE, f"Не удалось сохранить отчёт: {exc}")
            traceback.print_exc()

    def open_settings(self) -> None:
        global SETTINGS
        dialog = tk.Toplevel(self)
        dialog.title("Настройки правил")
        dialog.transient(self)
        dialog.grab_set()
        dialog.geometry("700x620")
        dialog.minsize(620, 520)

        ttk.Label(
            dialog,
            text="Отключённые автокоррекции останутся в проверке как ручные ошибки.",
            padding=(10, 10, 10, 4),
        ).pack(fill=tk.X)

        labels = {
            "check_duplicates": "Проверять дубли по всем файлам",
            "check_visible_sheet": "Строго проверять единственный видимый лист 'Шаблон'",
            "check_formulas": "Проверять затёртые формулы в расчетных колонках",
            "check_column_structure": "Проверять структуру колонок шире заголовков",
            "check_sheet_protection": "Проверять защиту листа 'Шаблон'",
            "parallel_scan": "Ускорять проверку параллельным чтением файлов",
            "autofix_sheet_protection": "Автоисправлять снятую защиту листа 'Шаблон'",
            "autofix_formulas": "Автоматически восстанавливать затёртые формулы",
            "autofix_column_structure": "Автоматически показывать скрытые колонки шаблона",
            "autofix_whitespace": "Автоисправлять пробелы и длинные тире",
            "autofix_reference_format": "Автоисправлять формат значений справочников и 'Нет'",
            "autofix_dates_years": "Автоисправлять даты, годы и расчетные сроки",
            "autofix_snils_format": "Автоформатировать корректный СНИЛС",
            "autofix_invalid_chars": "Автоисправлять безопасно заменяемые недопустимые символы",
            "autofix_conditional_cleanup": "Автоочищать поля, запрещенные условиями",
            "autofix_category_split": "Авторазделять списки категорий C, D, E, F на строки",
        }

        frame = ttk.Frame(dialog, padding=(10, 0, 10, 10))
        frame.pack(fill=tk.BOTH, expand=True)
        vars_by_name: dict[str, tk.BooleanVar] = {}
        for field_name, label in labels.items():
            var = tk.BooleanVar(value=getattr(SETTINGS, field_name))
            vars_by_name[field_name] = var
            ttk.Checkbutton(frame, text=label, variable=var).pack(anchor="w", pady=3)

        buttons = ttk.Frame(dialog, padding=10)
        buttons.pack(fill=tk.X)

        def save() -> None:
            global SETTINGS
            SETTINGS = RuleSettings(**{name: var.get() for name, var in vars_by_name.items()})
            save_settings(SETTINGS)
            dialog.destroy()
            self.scan()

        ttk.Button(buttons, text="Отмена", command=dialog.destroy).pack(side=tk.RIGHT)
        ttk.Button(buttons, text="Сохранить и проверить", command=save).pack(side=tk.RIGHT, padx=(0, 6))
        self.wait_window(dialog)

    def choose_files_for_fix(self, fixable: list[Issue]) -> set[Path] | None:
        by_file: dict[Path, list[Issue]] = {}
        for issue in fixable:
            by_file.setdefault(issue.file, []).append(issue)

        dialog = tk.Toplevel(self)
        dialog.title("Подтверждение исправлений")
        dialog.transient(self)
        dialog.grab_set()
        dialog.geometry("620x420")
        dialog.minsize(520, 320)

        ttk.Label(
            dialog,
            text="Выберите файлы, в которых нужно применить автоматические исправления:",
            padding=10,
        ).pack(fill=tk.X)

        outer = ttk.Frame(dialog, padding=(10, 0, 10, 10))
        outer.pack(fill=tk.BOTH, expand=True)
        canvas = tk.Canvas(outer, highlightthickness=0)
        scrollbar = ttk.Scrollbar(outer, orient=tk.VERTICAL, command=canvas.yview)
        content = ttk.Frame(canvas)
        content.bind("<Configure>", lambda _event: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=content, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        vars_by_file: dict[Path, tk.BooleanVar] = {}
        for path, file_issues in sorted(by_file.items(), key=lambda item: self.relative_file(item[0]).casefold()):
            var = tk.BooleanVar(value=True)
            vars_by_file[path] = var
            kinds = Counter(issue.message for issue in file_issues)
            preview = "; ".join(f"{message}: {count}" for message, count in kinds.most_common(3))
            if len(kinds) > 3:
                preview += f"; ещё типов: {len(kinds) - 3}"
            text_line = f"{self.relative_file(path)} - {len(file_issues)} исправл."
            ttk.Checkbutton(content, text=text_line, variable=var).pack(anchor="w", pady=(4, 0))
            ttk.Label(content, text=preview, foreground="#555555", wraplength=540).pack(anchor="w", padx=(24, 0))

        result: dict[str, Any] = {"files": None}

        buttons = ttk.Frame(dialog, padding=10)
        buttons.pack(fill=tk.X)

        def select_all(value: bool) -> None:
            for var in vars_by_file.values():
                var.set(value)

        def ok() -> None:
            result["files"] = {path for path, var in vars_by_file.items() if var.get()}
            dialog.destroy()

        def selected_report_items() -> dict[Path, list[Issue]]:
            return {
                path: by_file[path]
                for path, var in vars_by_file.items()
                if var.get()
            }

        def save_report() -> None:
            selected_items = selected_report_items()
            if not selected_items:
                messagebox.showinfo(APP_TITLE, "Не выбран ни один файл для отчёта.")
                return
            default_name = f"frdo_fix_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.doc"
            target = filedialog.asksaveasfilename(
                parent=dialog,
                title="Сохранить отчёт",
                defaultextension=".doc",
                initialdir=str(self.directory),
                initialfile=default_name,
                filetypes=(("Документ Word", "*.doc"), ("Все файлы", "*.*")),
            )
            if not target:
                return
            try:
                save_fix_report_doc(Path(target), selected_items, self.relative_file)
                messagebox.showinfo(APP_TITLE, f"Отчёт сохранён:\n{target}")
            except Exception as exc:
                messagebox.showerror(APP_TITLE, f"Не удалось сохранить отчёт: {exc}")
                traceback.print_exc()

        def cancel() -> None:
            dialog.destroy()

        ttk.Button(buttons, text="Все", command=lambda: select_all(True)).pack(side=tk.LEFT)
        ttk.Button(buttons, text="Ни одного", command=lambda: select_all(False)).pack(side=tk.LEFT, padx=(6, 0))
        ttk.Button(buttons, text="Сохранить отчёт .doc", command=save_report).pack(side=tk.LEFT, padx=(6, 0))
        ttk.Button(buttons, text="Отмена", command=cancel).pack(side=tk.RIGHT)
        ttk.Button(buttons, text="Применить", command=ok).pack(side=tk.RIGHT, padx=(0, 6))

        self.wait_window(dialog)
        return result["files"]

    def fix(self, issues: list[Issue], selected: bool) -> None:
        fixable = [issue for issue in issues if issue.proposed is not None]
        if not fixable:
            messagebox.showinfo(APP_TITLE, "Нет выбранных автоматически исправимых ошибок." if selected else "Нет автоматически исправимых ошибок.")
            return
        selected_files = self.choose_files_for_fix(fixable)
        if selected_files is None:
            return
        fixable = [issue for issue in fixable if issue.file in selected_files]
        if not fixable:
            messagebox.showinfo(APP_TITLE, "Не выбран ни один файл для исправления.")
            return
        before_total = len(self.issues)
        before_fixable = sum(1 for issue in self.issues if issue.proposed is not None)
        self.set_busy(True)
        try:
            saved = apply_fixes(fixable)
            self.scan()
            after_total = len(self.issues)
            after_fixable = sum(1 for issue in self.issues if issue.proposed is not None)
            messagebox.showinfo(
                APP_TITLE,
                "Исправления применены и выполнена повторная проверка:\n"
                + "\n".join(f"{p.name}: {count}" for p, count in saved.items())
                + f"\n\nПроблем всего: {before_total} -> {after_total}"
                + f"\nАвтоисправимых: {before_fixable} -> {after_fixable}",
            )
        except Exception as exc:
            messagebox.showerror(APP_TITLE, f"Ошибка при исправлении: {exc}")
            traceback.print_exc()
        finally:
            self.set_busy(False)


def cli_scan(directory: Path) -> int:
    files = [path for path in sorted(directory.rglob("*.xlsx")) if not path.name.startswith("~$")]
    issues = scan_files(files)
    if SETTINGS.check_duplicates:
        issues.extend(duplicate_issues_for_files(files))
    issues = deduplicate_issues(issues)
    fixable = sum(1 for issue in issues if issue.proposed is not None)
    print(f"Файлов проверено: {len(files)}")
    print(f"Проблем найдено: {len(issues)}")
    print(f"Можно исправить автоматически: {fixable}")
    for issue in issues[:200]:
        display_file = str(issue.file.relative_to(directory)) if issue.file.is_relative_to(directory) else issue.file.name
        print(
            f"{display_file}; {issue.address}; {issue.field}; {issue.status}; "
            f"{issue.message}; сейчас={value_for_display(issue.current)!r}; исправить={value_for_display(issue.proposed)!r}"
        )
    if len(issues) > 200:
        print(f"... ещё {len(issues) - 200} проблем")
    return 1 if issues else 0


def main() -> None:
    if "--cli" in sys.argv:
        directory = Path(sys.argv[sys.argv.index("--cli") + 1]) if len(sys.argv) > sys.argv.index("--cli") + 1 else Path.cwd()
        raise SystemExit(cli_scan(directory))
    app = CheckerApp()
    app.mainloop()


if __name__ == "__main__":
    main()
