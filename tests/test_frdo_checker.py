from __future__ import annotations

import tempfile
import unittest
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

import frdo_spo_checker as checker


class FrdoCheckerTests(unittest.TestCase):
    def test_settings_roundtrip(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / checker.SETTINGS_FILE
            settings = checker.RuleSettings(check_duplicates=False, autofix_category_split=False)
            with patch.object(checker, "settings_path", return_value=path):
                checker.save_settings(settings)
                loaded = checker.load_settings()
        self.assertFalse(loaded.check_duplicates)
        self.assertFalse(loaded.autofix_category_split)
        self.assertTrue(loaded.autofix_dates_years)

    def test_manual_snils_issue_has_hint(self) -> None:
        issue = checker.Issue(Path("sample.xlsx"), 2, 1, "СНИЛС", "Неверный СНИЛС")
        self.assertIn("000-000-000 00", issue.hint)

    def test_invalid_chars_autofix_can_be_disabled(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.cell(2, checker.COL["Номер документа"]).value = "12 A"
        issues: list[checker.Issue] = []
        old_settings = checker.SETTINGS
        try:
            checker.SETTINGS = checker.RuleSettings(autofix_invalid_chars=False)
            checker.check_regex(
                issues,
                Path("sample.xlsx"),
                ws,
                2,
                "Номер документа",
                r"\d+",
                "Номер должен содержать только цифры",
                checker.sanitize_digits,
            )
        finally:
            checker.SETTINGS = old_settings
        self.assertEqual(len(issues), 1)
        self.assertIsNone(issues[0].proposed)

    def test_category_split_copies_rows_by_blocks(self) -> None:
        wb = Workbook()
        ws = wb.active
        for col, header in enumerate(checker.PO_HEADERS, start=1):
            ws.cell(1, col).value = header
        category_col = checker.po_col("Присвоенный квалификационный разряд, класс, категория (при наличии)")
        ws.cell(2, 1).value = "row-2"
        ws.cell(3, 1).value = "row-3"
        ws.cell(2, category_col).value = "C,D,E"
        ws.cell(3, category_col).value = "C,D,E"
        issues = [
            checker.Issue(Path("po.xlsx"), 2, category_col, "Категория", "split", "C,D,E", checker.CategorySplit(("C", "D", "E")), "ПО"),
            checker.Issue(Path("po.xlsx"), 3, category_col, "Категория", "split", "C,D,E", checker.CategorySplit(("C", "D", "E")), "ПО"),
        ]

        checker.apply_category_splits(ws, issues)

        self.assertEqual([ws.cell(row, category_col).value for row in range(2, 8)], ["C", "C", "D", "D", "E", "E"])
        self.assertEqual([ws.cell(row, 1).value for row in range(2, 8)], ["row-2", "row-3", "row-2", "row-3", "row-2", "row-3"])

    def test_category_split_accepts_any_driver_category_combination(self) -> None:
        for raw, expected in {
            "C,E,F": ("C", "E", "F"),
            "C,F,D": ("C", "F", "D"),
            "E; D": ("E", "D"),
        }.items():
            with self.subTest(raw=raw):
                ok, split = checker.valid_category_list(raw, [])

                self.assertTrue(ok)
                self.assertEqual(split, checker.CategorySplit(expected))

    def test_category_split_preserves_valid_style_references(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "po.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = checker.SHEET_NAME
            for col, header in enumerate(checker.PO_HEADERS, start=1):
                ws.cell(1, col).value = header
            category_col = checker.po_col("Присвоенный квалификационный разряд, класс, категория (при наличии)")
            styled_cell = ws.cell(2, checker.po_col("Номер документа"))
            styled_cell.value = "1"
            styled_cell.font = Font(name="Arial", bold=True, color="FF000100")
            styled_cell.fill = PatternFill("solid", fgColor="FFFF0100")
            ws.cell(2, checker.po_col("Серия документа")).value = "Нет"
            ws.cell(2, category_col).value = "C,E,F"
            wb.save(path)
            wb.close()

            checker.apply_fixes([
                checker.Issue(
                    path,
                    2,
                    category_col,
                    "Категория",
                    "split",
                    "C,E,F",
                    checker.CategorySplit(("C", "E", "F")),
                    "ПО",
                )
            ])

            with zipfile.ZipFile(path) as archive:
                styles_xml = archive.read("xl/styles.xml")
                sheet_xml = archive.read("xl/worksheets/sheet1.xml")
            ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            styles_root = ET.fromstring(styles_xml)
            sheet_root = ET.fromstring(sheet_xml)
            cell_xfs = styles_root.find("main:cellXfs", ns)
            self.assertIsNotNone(cell_xfs)
            style_count = int(cell_xfs.attrib["count"])
            style_ids = [int(cell.attrib["s"]) for cell in sheet_root.findall(".//main:c[@s]", ns)]
            self.assertTrue(style_ids)
            self.assertLess(max(style_ids), style_count)

    def test_formula_and_protection_checks(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = checker.SHEET_NAME
        for col, header in enumerate(checker.EXPECTED_HEADERS, start=1):
            ws.cell(1, col).value = header
        ws.protection.sheet = False
        ws.cell(2, checker.COL["Серия документа"]).value = "Нет"
        ws.cell(2, checker.COL["Номер документа"]).value = "1"
        ws.cell(2, checker.COL["Год поступления"]).value = 2020
        ws.cell(2, checker.COL["Год окончания"]).value = 2024
        ws.cell(2, checker.COL["Срок обучения, лет"]).value = "=Q2-P2"
        ws.cell(3, checker.COL["Серия документа"]).value = "Нет"
        ws.cell(3, checker.COL["Номер документа"]).value = "2"
        ws.cell(3, checker.COL["Год поступления"]).value = 2020
        ws.cell(3, checker.COL["Год окончания"]).value = 2024
        ws.cell(3, checker.COL["Срок обучения, лет"]).value = 4
        issues: list[checker.Issue] = []

        old_settings = checker.SETTINGS
        try:
            checker.SETTINGS = checker.RuleSettings(check_formulas=True, check_sheet_protection=True, autofix_sheet_protection=True)
            checker.check_sheet_protection(issues, Path("sample.xlsx"), ws, "СПО")
            checker.check_formula_columns(issues, Path("sample.xlsx"), ws, checker.EXPECTED_HEADERS, checker.row_has_data, "СПО")
        finally:
            checker.SETTINGS = old_settings

        messages = [issue.message for issue in issues]
        self.assertIn("Защита листа 'Шаблон' снята", messages)
        self.assertIn("Формула в расчетной колонке отсутствует или заменена значением", messages)
        protection_issue = next(issue for issue in issues if issue.action == "protect_sheet")
        self.assertEqual(protection_issue.proposed, "включить защиту листа")
        formula_issue = next(issue for issue in issues if "Формула" in issue.message)
        self.assertEqual(formula_issue.proposed, "=Q3-P3")

    def test_validation_structure_is_checked_only_for_reference_fields(self) -> None:
        wb = Workbook()
        ws = wb.active
        for col, header in enumerate(checker.EXPECTED_HEADERS, start=1):
            ws.cell(1, col).value = header
        snils_col = 24
        doc_type_col = 2
        for row in (2, 3):
            ws.cell(row, snils_col).value = "123-456-789 00"
            ws.cell(row, doc_type_col).value = "value"
        snils_validation = DataValidation(type="textLength", operator="equal", formula1="14")
        doc_type_validation = DataValidation(type="list", formula1='"value"')
        ws.add_data_validation(snils_validation)
        ws.add_data_validation(doc_type_validation)
        snils_validation.add(ws.cell(2, snils_col))
        doc_type_validation.add(ws.cell(2, doc_type_col))

        issues: list[checker.Issue] = []
        checker.check_column_structure(
            issues,
            Path("sample.xlsx"),
            ws,
            checker.EXPECTED_HEADERS,
            lambda sheet, row: True,
            "РЎРџРћ",
            checker.SPO_VALIDATION_FIELDS,
        )

        issue_fields = [issue.field for issue in issues]
        self.assertIn(checker.EXPECTED_HEADERS[1], issue_fields)
        self.assertNotIn(checker.EXPECTED_HEADERS[23], issue_fields)

    def test_spo_extra_matrix_rules_for_dates_and_target_fields(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = checker.SHEET_NAME
        for col, header in enumerate(checker.EXPECTED_HEADERS, start=1):
            ws.cell(1, col).value = header
        ws.cell(2, 8).value = "Нет"
        ws.cell(2, 9).value = "1"
        ws.cell(2, 10).value = "01.01.2024"
        ws.cell(2, 29).value = "Да"
        ws.cell(2, 30).value = "12@"
        ws.cell(2, 31).value = "02.01.2024"
        ws.cell(2, 32).value = "ООО @"
        ws.cell(2, 35).value = "ООО @"
        ws.cell(2, 43).value = "03.01.2024"

        issues = checker.scan_spo_workbook(Path("sample.xlsx"), wb)
        messages = [issue.message for issue in issues]

        self.assertIn("Дата договора о целевом обучении позже даты выдачи документа", messages)
        self.assertIn("Дата выдачи оригинала позже даты выдачи дубликата", messages)
        self.assertIn("Недопустимые символы в номере договора о целевом обучении", messages)
        self.assertIn("Недопустимые символы в наименовании организации", messages)

    def test_spo_issue_date_year_must_not_be_before_graduation_year(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = checker.SHEET_NAME
        for col, header in enumerate(checker.EXPECTED_HEADERS, start=1):
            ws.cell(1, col).value = header
        ws.cell(2, checker.COL["Серия документа"]).value = "Нет"
        ws.cell(2, checker.COL["Номер документа"]).value = "1"
        ws.cell(2, checker.COL["Дата выдачи"]).value = "01.06.2023"
        ws.cell(2, checker.COL["Год поступления"]).value = 2020
        ws.cell(2, checker.COL["Год окончания"]).value = 2024
        ws.cell(2, checker.COL["Срок обучения, лет"]).value = 4

        issues = checker.scan_spo_workbook(Path("sample.xlsx"), wb)
        issue = next(item for item in issues if item.message == "Год даты выдачи меньше года окончания")

        self.assertEqual(issue.col, checker.COL["Дата выдачи"])
        self.assertEqual(issue.severity, checker.SEVERITY_BLOCKING)

    def test_po_extra_matrix_rules_for_original_document(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = checker.SHEET_NAME
        for col, header in enumerate(checker.PO_HEADERS, start=1):
            ws.cell(1, col).value = header
        ws.cell(2, 1).value = "Свидетельство"
        ws.cell(2, 2).value = "Дубликат"
        ws.cell(2, 6).value = "Нет"
        ws.cell(2, 7).value = "1"
        ws.cell(2, 8).value = "01.01.2024"
        ws.cell(2, 27).value = "Свидетельство"
        ws.cell(2, 29).value = "12A"
        ws.cell(2, 31).value = "02.01.2024"

        issues = checker.scan_po_workbook(Path("sample.xlsx"), wb)
        messages = [issue.message for issue in issues]

        self.assertIn("Номер оригинала должен содержать только цифры", messages)
        self.assertIn("Дата выдачи оригинала позже даты выдачи дубликата", messages)

    def test_po_issue_date_year_must_not_be_before_graduation_year(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = checker.SHEET_NAME
        for col, header in enumerate(checker.PO_HEADERS, start=1):
            ws.cell(1, col).value = header
        ws.cell(2, checker.po_col("Серия документа")).value = "Нет"
        ws.cell(2, checker.po_col("Номер документа")).value = "1"
        ws.cell(2, checker.po_col("Дата выдачи документа")).value = "01.06.2023"
        ws.cell(2, checker.po_col("Год начала обучения")).value = 2020
        ws.cell(2, checker.po_col("Год окончания обучения")).value = 2024
        ws.cell(2, checker.po_col("Срок обучения, часов")).value = 72

        issues = checker.scan_po_workbook(Path("sample.xlsx"), wb)
        issue = next(item for item in issues if item.message == "Год даты выдачи меньше года окончания обучения")

        self.assertEqual(issue.col, checker.po_col("Дата выдачи документа"))
        self.assertEqual(issue.severity, checker.SEVERITY_BLOCKING)

    def test_manual_edit_creates_backup_and_writes_value(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "sample.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = checker.SHEET_NAME
            ws.cell(1, 1).value = checker.EXPECTED_HEADERS[0]
            ws.cell(2, 1).value = "старое"
            wb.save(path)
            wb.close()

            issue = checker.Issue(path, 2, 1, checker.EXPECTED_HEADERS[0], "ручная правка", "старое", profile="СПО")
            backup = checker.apply_manual_edit(issue, "новое")

            self.assertTrue(backup.exists())
            self.assertEqual(backup.parent.name, "_bak")
            loaded = checker.load_workbook(path, read_only=False, data_only=False)
            try:
                self.assertEqual(loaded[checker.SHEET_NAME].cell(2, 1).value, "новое")
            finally:
                loaded.close()

    def test_apply_fixes_enables_sheet_protection(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "sample.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = checker.SHEET_NAME
            ws.protection.sheet = False
            wb.save(path)
            wb.close()

            issue = checker.Issue(
                path,
                0,
                0,
                "",
                "Защита листа 'Шаблон' снята",
                "защита отключена",
                "включить защиту листа",
                "СПО",
                "protect_sheet",
            )
            checker.apply_fixes([issue])

            loaded = checker.load_workbook(path, read_only=False, data_only=False)
            try:
                self.assertTrue(loaded[checker.SHEET_NAME].protection.sheet)
            finally:
                loaded.close()

    def test_apply_fixes_restores_formula_and_unhides_column(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "sample.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = checker.SHEET_NAME
            ws.cell(1, checker.COL["Срок обучения, лет"]).value = "Срок обучения, лет"
            ws.cell(3, checker.COL["Срок обучения, лет"]).value = 4
            ws.column_dimensions["R"].hidden = True
            wb.save(path)
            wb.close()

            issues = [
                checker.Issue(path, 3, checker.COL["Срок обучения, лет"], "Срок обучения, лет", "Формула", 4, "=Q3-P3", "СПО"),
                checker.Issue(path, 1, checker.COL["Срок обучения, лет"], "Срок обучения, лет", "Колонка шаблона скрыта", "скрыта", "показать колонку", "СПО", "unhide_column"),
            ]
            checker.apply_fixes(issues)

            loaded = checker.load_workbook(path, read_only=False, data_only=False)
            try:
                ws_loaded = loaded[checker.SHEET_NAME]
                self.assertEqual(ws_loaded.cell(3, checker.COL["Срок обучения, лет"]).value, "=Q3-P3")
                self.assertFalse(ws_loaded.column_dimensions["R"].hidden)
            finally:
                loaded.close()

    def test_manual_value_validation(self) -> None:
        date_issue = checker.Issue(Path("sample.xlsx"), 2, checker.COL["Дата выдачи"], "Дата выдачи", "ручная дата")
        ok, _error, value = checker.validate_manual_value(date_issue, "01.02.2024")
        self.assertTrue(ok)
        self.assertEqual(value.strftime("%d.%m.%Y"), "01.02.2024")

        snils_issue = checker.Issue(Path("sample.xlsx"), 2, checker.COL["СНИЛС"], "СНИЛС", "ручной СНИЛС")
        ok, error, _value = checker.validate_manual_value(snils_issue, "123")
        self.assertFalse(ok)
        self.assertIn("СНИЛС", error)

    def test_fix_report_doc_is_created(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            report = Path(tmp) / "report.doc"
            file_path = Path(tmp) / "sample.xlsx"
            issue = checker.Issue(file_path, 2, 1, "Серия документа", "Серия документа не заполнена", "", "Нет", "СПО")

            checker.save_fix_report_doc(report, {file_path: [issue]}, lambda path: path.name)

            content = report.read_text(encoding="utf-8")
            self.assertIn("sample.xlsx", content)
            self.assertIn("Серия документа не заполнена", content)
            self.assertIn("Всего исправлений: 1", content)

    def test_issue_table_report_doc_is_created(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            report = Path(tmp) / "table_report.doc"
            issue = checker.Issue(
                Path(tmp) / "sample.xlsx",
                2,
                1,
                "Серия документа",
                "Серия документа не заполнена",
                "",
                "Нет",
                "СПО",
            )

            checker.save_issue_table_report_doc(report, [issue], lambda path: path.name)

            content = report.read_text(encoding="utf-8")
            self.assertIn("Отчёт по таблице ошибок ФРДО", content)
            self.assertIn("sample.xlsx", content)
            self.assertIn("Серия документа", content)
            self.assertIn("Строк в отчёте: 1", content)
            self.assertIn("mso-page-orientation: landscape", content)
            self.assertIn("margin: 0.7cm", content)
            self.assertIn("table-layout: fixed", content)
            self.assertIn("mso-table-header-row", content)
            self.assertNotIn("<th>Подсказка</th>", content)
            self.assertNotIn("<th>Исправить на</th>", content)

    def test_merge_spo_workbooks_copies_rows_and_writes_calculated_term(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            first = root / "spo_1.xlsx"
            second = root / "spo_2.xlsx"
            target = root / "merged.xlsx"

            for path, doc_number in ((first, "1"), (second, "2")):
                wb = Workbook()
                ws = wb.active
                ws.title = checker.SHEET_NAME
                for col, header in enumerate(checker.EXPECTED_HEADERS, start=1):
                    ws.cell(1, col).value = header
                ws.cell(2, checker.COL["Серия документа"]).value = "Нет"
                ws.cell(2, checker.COL["Номер документа"]).value = doc_number
                ws.cell(2, checker.COL["Год поступления"]).value = 2020
                ws.cell(2, checker.COL["Год окончания"]).value = 2024
                ws.cell(2, checker.COL["Срок обучения, лет"]).value = "=Q2-P2"
                wb.save(path)
                wb.close()

            rows = checker.merge_profile_workbooks([first, second], checker.PROFILES["spo"], target)

            self.assertEqual(rows, 2)
            loaded = checker.load_workbook(target, read_only=False, data_only=False)
            try:
                ws = loaded[checker.SHEET_NAME]
                self.assertEqual(ws.cell(2, checker.COL["Номер документа"]).value, "1")
                self.assertEqual(ws.cell(3, checker.COL["Номер документа"]).value, "2")
                self.assertEqual(ws.cell(2, checker.COL["Срок обучения, лет"]).value, 4)
                self.assertEqual(ws.cell(3, checker.COL["Срок обучения, лет"]).value, 4)
            finally:
                loaded.close()

    def test_merge_workbooks_keeps_style_references_valid(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            first = root / "spo_1.xlsx"
            second = root / "spo_2.xlsx"
            target = root / "merged.xlsx"

            for index, path in enumerate((first, second), start=1):
                wb = Workbook()
                ws = wb.active
                ws.title = checker.SHEET_NAME
                for col, header in enumerate(checker.EXPECTED_HEADERS, start=1):
                    ws.cell(1, col).value = header
                cell = ws.cell(2, checker.COL["Номер документа"])
                cell.value = str(index)
                cell.font = Font(name="Arial", bold=True, color=f"FF000{index}00")
                cell.fill = PatternFill("solid", fgColor=f"FFFF0{index}00")
                ws.cell(2, checker.COL["Серия документа"]).value = "Нет"
                ws.cell(2, checker.COL["Год поступления"]).value = 2020
                ws.cell(2, checker.COL["Год окончания"]).value = 2024
                wb.save(path)
                wb.close()

            checker.merge_profile_workbooks([first, second], checker.PROFILES["spo"], target)

            with zipfile.ZipFile(target) as archive:
                styles_xml = archive.read("xl/styles.xml")
                sheet_xml = archive.read("xl/worksheets/sheet1.xml")
            ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            styles_root = ET.fromstring(styles_xml)
            sheet_root = ET.fromstring(sheet_xml)
            cell_xfs = styles_root.find("main:cellXfs", ns)
            self.assertIsNotNone(cell_xfs)
            style_count = int(cell_xfs.attrib["count"])
            style_ids = [
                int(cell.attrib["s"])
                for cell in sheet_root.findall(".//main:c[@s]", ns)
            ]
            self.assertTrue(style_ids)
            self.assertLess(max(style_ids), style_count)

    def test_mergeable_files_for_profile_skips_only_blocking_files(self) -> None:
        clean = Path("clean.xlsx")
        blocked = Path("blocked.xlsx")
        files = [clean, blocked]
        issues = [
            checker.Issue(
                clean,
                1,
                checker.COL["Срок обучения, лет"],
                "Срок обучения, лет",
                "Подозрительная ширина колонки: 100",
                profile="СПО",
            ),
            checker.Issue(
                blocked,
                2,
                checker.COL["Номер документа"],
                "Номер документа",
                "Обязательное поле не заполнено",
                profile="СПО",
            )
        ]

        merge_files, blocking_by_file = checker.mergeable_files_for_profile(files, checker.PROFILES["spo"], issues)

        self.assertEqual(merge_files, [clean])
        self.assertEqual(blocking_by_file, {blocked: 1})

    def test_xlsx_scan_skips_generated_merge_folder(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source.xlsx"
            generated_in_root = root / "frdo_merged_spo_20260701_120000.xlsx"
            generated_dir = root / "_merged"
            generated_dir.mkdir()
            generated = generated_dir / "frdo_merged_spo.xlsx"
            generated_double_dir = root / "__merged"
            generated_double_dir.mkdir()
            generated_double = generated_double_dir / "manual_merged.xlsx"
            source.touch()
            generated_in_root.touch()
            generated.touch()
            generated_double.touch()

            files = checker.xlsx_files_for_scan(root)

            self.assertIn(source, files)
            self.assertNotIn(generated_in_root, files)
            self.assertNotIn(generated, files)
            self.assertNotIn(generated_double, files)


if __name__ == "__main__":
    unittest.main()
